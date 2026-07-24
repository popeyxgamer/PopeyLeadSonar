# -*- coding: utf-8 -*-
import csv
from PySide6.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QComboBox,
    QPushButton, QTableWidget, QTableWidgetItem, QMessageBox, QFileDialog, QScrollArea
)
from PySide6.QtGui import QColor
from PySide6.QtCore import Qt
from ui.views.base_view import BaseView
from ui.i18n import tr
from ui.styles import COLOR_OK, COLOR_ERROR
from core import database as db
from core.signal_bus import bus

class LeadsView(BaseView):
    def setup_ui(self):
        header = QLabel(tr("Zarządzanie Leadami"))
        header.setStyleSheet("font-size: 20px; font-weight: bold; color: white;")
        self.layout.addWidget(header)

        # Toolbar
        row = QHBoxLayout()
        self.filter_status = QComboBox()
        self.filter_status.addItems(["Wszystkie", "nowy", "wysłano", "błąd"])
        self.filter_status.currentTextChanged.connect(self._on_filter_changed)
        row.addWidget(QLabel(tr('Status:')))
        row.addWidget(self.filter_status)

        self.filter_search = QLineEdit()
        self.filter_search.setPlaceholderText(tr('Szukaj firmy lub emaila...'))
        self.filter_search.textChanged.connect(self._on_filter_changed)
        row.addWidget(self.filter_search)

        btn_refresh = QPushButton(tr('🔄 Odśwież'))
        btn_refresh.clicked.connect(self.refresh_leads)
        row.addWidget(btn_refresh)

        btn_delete_sent = QPushButton(tr('🗑 Usuń wysłane'))
        btn_delete_sent.clicked.connect(self.delete_sent)
        row.addWidget(btn_delete_sent)

        row.addStretch()
        self.layout.addLayout(row)

        # Pagination Toolbar
        pag_row = QHBoxLayout()
        self.btn_prev = QPushButton("← " + tr("Poprzednia"))
        self.btn_prev.clicked.connect(self.prev_page)
        pag_row.addWidget(self.btn_prev)

        self.page_label = QLabel(tr("Strona 1"))
        self.page_label.setStyleSheet("font-weight: bold; padding: 0 10px;")
        pag_row.addWidget(self.page_label)

        self.btn_next = QPushButton(tr("Następna") + " →")
        self.btn_next.clicked.connect(self.next_page)
        pag_row.addWidget(self.btn_next)

        pag_row.addStretch()
        self.layout.addLayout(pag_row)

        # Table
        self.leads_table = QTableWidget()
        self.leads_table.setColumnCount(7)
        self.leads_table.setHorizontalHeaderLabels(
            [tr('ID'), tr('Firma'), tr('Email'), tr('Adres'), tr('Typ'), tr('Status'), tr('Wysłano')]
        )
        self.leads_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.leads_table.horizontalHeader().setStretchLastSection(True)
        self.layout.addWidget(self.leads_table)

        # Footer Actions
        row2 = QHBoxLayout()
        btn_export = QPushButton(tr('📤 Eksportuj CSV'))
        btn_export.clicked.connect(self.export_csv)
        row2.addWidget(btn_export)

        btn_import = QPushButton(tr('📥 Importuj CSV/Excel'))
        btn_import.clicked.connect(self.import_csv_excel)
        row2.addWidget(btn_import)

        btn_blacklist = QPushButton(tr('🚫 Dodaj do czarnej listy'))
        btn_blacklist.clicked.connect(self.add_selected_to_blacklist)
        row2.addWidget(btn_blacklist)

        row2.addStretch()
        self.layout.addLayout(row2)

        # State
        self.current_page = 0
        self.page_size = 100
        self.total_leads = 0

    def setup_signals(self):
        bus.leads_changed.connect(self.refresh_leads)
        bus.profile_changed.connect(self.reset_and_refresh)

    def reset_and_refresh(self):
        self.current_page = 0
        self.refresh_leads()

    def _on_filter_changed(self):
        self.current_page = 0
        self.refresh_leads()

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.refresh_leads()

    def next_page(self):
        if (self.current_page + 1) * self.page_size < self.total_leads:
            self.current_page += 1
            self.refresh_leads()

    def refresh_leads(self):
        status = self.filter_status.currentText()
        status = None if status == "Wszystkie" else status
        search = self.filter_search.text().strip().lower()

        self.total_leads = db.count_leads(status=status, search=search)
        rows = db.get_leads_summary(
            status=status, search=search,
            limit=self.page_size, offset=self.current_page * self.page_size
        )

        # Update UI state
        self.page_label.setText(tr("Strona {} z {}").format(
            self.current_page + 1, max(1, (self.total_leads + self.page_size - 1) // self.page_size)
        ))
        self.btn_prev.setEnabled(self.current_page > 0)
        self.btn_next.setEnabled((self.current_page + 1) * self.page_size < self.total_leads)

        self.leads_table.setRowCount(len(rows))
        STATUS_COL = 5
        for i, row in enumerate(rows):
            for col, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val else "")
                if col == STATUS_COL:
                    if val == 'wysłano':
                        item.setBackground(QColor(*COLOR_OK))
                    elif val == 'błąd':
                        item.setBackground(QColor(*COLOR_ERROR))
                self.leads_table.setItem(i, col, item)
        self.leads_table.resizeColumnsToContents()

    def setup_signals(self):
        bus.leads_changed.connect(self.refresh_leads)
        bus.profile_changed.connect(self.refresh_leads)

    def refresh_leads(self):
        status = self.filter_status.currentText()
        status = None if status == "Wszystkie" else status
        search = self.filter_search.text().strip().lower()
        rows = db.get_leads_summary(status=status, search=search)

        self.leads_table.setRowCount(len(rows))
        STATUS_COL = 5
        for i, row in enumerate(rows):
            for col, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val else "")
                if col == STATUS_COL:
                    if val == 'wysłano':
                        item.setBackground(QColor(*COLOR_OK))
                    elif val == 'błąd':
                        item.setBackground(QColor(*COLOR_ERROR))
                self.leads_table.setItem(i, col, item)
        self.leads_table.resizeColumnsToContents()

    def delete_sent(self):
        reply = QMessageBox.question(
            self, tr('Potwierdzenie'), tr('Usunąć wszystkie wysłane leady?'),
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            db.delete_sent_leads()
            bus.leads_changed.emit()

    def export_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, "Zapisz CSV", "leady.csv", "CSV Files (*.csv)")
        if not path:
            return
        rows = db.get_leads()
        headers = ["id", "firma", "kontakt", "email", "adres", "telefon", "website", "typ", "status", "wyslano"]
        try:
            with open(path, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(rows)
            QMessageBox.information(self, tr('OK'), f"Eksportowano {len(rows)} leadów")
        except OSError as e:
            QMessageBox.critical(self, tr('Błąd'), f"Nie można zapisać pliku: {e}")

    def import_csv_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Wybierz plik", "",
            "Pliki CSV/Excel (*.csv *.xlsx *.xls)"
        )
        if not path:
            return
        try:
            imported = 0
            if path.endswith('.csv'):
                with open(path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        email = (row.get('email') or '').strip()
                        if email:
                            if db.add_lead(
                                row.get('firma', ''), email, row.get('adres', ''),
                                row.get('telefon', ''), row.get('website', ''),
                                row.get('typ', ''), row.get('kontakt', '')
                            ):
                                imported += 1
            else:
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(path, read_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    col_map = {h.lower(): i for i, h in enumerate(headers)}
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        email = row[col_map.get('email', -1)] if 'email' in col_map else None
                        if email:
                            firma = row[col_map.get('firma', -1)] if 'firma' in col_map else ''
                            adres = row[col_map.get('adres', -1)] if 'adres' in col_map else ''
                            telefon = row[col_map.get('telefon', -1)] if 'telefon' in col_map else ''
                            website = row[col_map.get('website', -1)] if 'website' in col_map else ''
                            typ = row[col_map.get('typ', -1)] if 'typ' in col_map else ''
                            kontakt = row[col_map.get('kontakt', -1)] if 'kontakt' in col_map else ''
                            if db.add_lead(firma, email, adres, telefon, website, typ, kontakt):
                                imported += 1
                except ImportError:
                    QMessageBox.warning(self, tr('Błąd'), tr('Zainstaluj openpyxl: pip install openpyxl'))
                    return

            bus.leads_changed.emit()
            QMessageBox.information(self, tr('OK'), f"Zaimportowano {imported} leadów")
        except Exception as e:
            QMessageBox.critical(self, tr('Błąd'), f"Nie można zaimportować: {e}")

    def add_selected_to_blacklist(self):
        selected_rows = set()
        for item in self.leads_table.selectedItems():
            selected_rows.add(item.row())
        count = 0
        for row in selected_rows:
            email_item = self.leads_table.item(row, 2)
            if email_item and email_item.text():
                if db.add_to_blacklist(email_item.text(), "manual"):
                    count += 1
        QMessageBox.information(self, tr('OK'), f"Dodano {count} adresów do blacklist.")
