# -*- coding: utf-8 -*-
"""Główne okno aplikacji – z przełącznikiem profili i zarządzaniem nimi."""
import csv
import os
from typing import Optional, List, Dict, Any

from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor, QAction
from PySide6.QtWidgets import (
    QCheckBox, QComboBox, QDialog, QDoubleSpinBox, QFileDialog, QFormLayout, QGroupBox,
    QHBoxLayout, QInputDialog, QLabel, QLineEdit, QListWidget, QListWidgetItem,
    QMainWindow, QMessageBox, QProgressBar, QPushButton, QSpinBox, QTableWidget,
    QTableWidgetItem, QTabWidget, QTextEdit, QVBoxLayout, QWidget, QScrollArea,
    QToolBar, QMenu, QMenuBar, QWizard,
)

from core import database as db
from core.ai_providers import ai_manager, OpenAIProvider, GeminiProvider, OllamaProvider, LMStudioProvider, DeepSeekLaudeProvider
from core.ai_features import TemplateGenerator, SubjectLineOptimizer, LeadPersonalizer, LeadScorer, ResponseAnalyzer, SendTimingOptimizer, ABTestingEngine
from core.ai_workers import AIWorker, BatchAIWorker
from core.config import (
    CUSTOM_SEND_DELAY_DEFAULT, CUSTOM_SEND_DELAY_MIN, CUSTOM_SESSION_CAP_DEFAULT,
    CUSTOM_SESSION_CAP_MAX, GMAIL_FREE_SESSION_CAP_DEFAULT, GMAIL_FREE_SESSION_CAP_OPTIONS,
    SEND_FIXED_DELAY, SESSION_CAP_OPTIONS, SESSION_HARD_CAP, SMTP_FALLBACK_HOST,
    SMTP_FALLBACK_PORT, SMTP_RELAY_HOST, SMTP_RELAY_PORT, get_abs_session_cap,
    get_send_delay, logger, guess_smtp,
)
from core.profile_manager import (
    get_all_profiles, get_current_profile_name, get_current_profile_settings,
    update_current_profile_settings, switch_profile, create_new_profile,
    delete_profile_by_name, copy_profile, get_profile_path, get_company_info,
)
from core.email_sender import test_gmail_connection, wyslij_email
from core.workers import (
    AutoPilotWorker, SearchWorker, SendWorker, InboxFetchWorker, MessageBodyWorker,
    FolderListWorker, MessageFullWorker, MessageActionWorker, AIAutoSendWorker,
)
from core.mailbox_reader import guess_imap_server, guess_trash_folder
from core.account_rotator import SMTPAccountRotator
from core.bounce_imap import BounceMonitor
from core.blacklist_import import import_blacklist_from_file
from ui.styles import (
    COLOR_ERROR, COLOR_OK, COLOR_SENT_LIST, COLOR_SENT_LIST_ALT,
    DARK_STYLESHEET, DEFAULT_LOCATIONS, DEFAULT_PROFILE_NAME, DEFAULT_QUERIES,
    DEFAULT_SUBJECT, DEFAULT_TEMPLATE,
)
from ui.profile_wizard import ProfileWizard
from ui.quickstart_wizard import QuickStartWizard
from ui.i18n import tr, get_language, set_language, load_language_from_disk, restart_app, SUPPORTED_LANGUAGES, LANGUAGE_NAMES, LANGUAGE_FLAGS

# Matplotlib
try:
    import matplotlib
    matplotlib.use('Qt5Agg')
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False


class MainWindow(QMainWindow):
    # Sygnał do odświeżenia UI po zmianie profilu
    profile_changed = Signal(str)

    def __init__(self):
        super().__init__()
        self.setWindowTitle('PopeyLeadSonar – ' + tr('🎯 Hybrydowy System Pozyskiwania i Wysyłki Leadów'))
        self.setWindowState(Qt.WindowMaximized)

        self.search_worker = None
        self.send_worker = None
        self.autopilot_worker = None
        self.bounce_monitor = None
        self.found_leads = []
        self.accounts_data = []  # lista kont do rotacji
        self._current_profile = get_current_profile_name()

        self.setStyleSheet(DARK_STYLESHEET)

        # Główny widget i layout
        self.central = QWidget()
        self.setCentralWidget(self.central)
        self.layout = QVBoxLayout(self.central)
        self.layout.setContentsMargins(0, 0, 0, 0)

        # Pasek narzędzi z przełącznikiem profili
        self._create_toolbar()

        # ------------------------------------------------------------
        # ZAKŁADKI GŁÓWNE – uporządkowane w 3 logiczne grupy:
        #   1) Kampania   -> szukanie firm, lista leadów, historia, asystent AI
        #   2) Wysyłka    -> wysyłka maili, skrzynka odbiorcza
        #   3) Ustawienia -> konfiguracja (podstawowe / AI / zaawansowane)
        # Dzięki temu aplikacja jest czytelna, a AI jest opcjonalnym
        # dodatkiem – wszystko działa też bez skonfigurowanego AI.
        # ------------------------------------------------------------
        self.tabs = QTabWidget()
        self.tabs.setObjectName("MainTabs")
        self.layout.addWidget(self.tabs)

        # -- Widgety "liściaste" (zawartość poszczególnych podzakładek) --
        self.search_tab = QWidget()
        self.leads_tab = QWidget()
        self.send_tab = QWidget()
        self.settings_tab = QWidget()
        self.history_tab = QWidget()
        self.inbox_tab = QWidget()
        self.ai_tab = QWidget()

        # Layouty dla zakładek (potrzebne do scrolla)
        self.search_tab.setLayout(QVBoxLayout())
        self.leads_tab.setLayout(QVBoxLayout())
        self.send_tab.setLayout(QVBoxLayout())
        self.settings_tab.setLayout(QVBoxLayout())
        self.history_tab.setLayout(QVBoxLayout())
        self.inbox_tab.setLayout(QVBoxLayout())
        self.ai_tab.setLayout(QVBoxLayout())

        # -- Grupa 1: KAMPANIA (szukanie + lista + historia + AI) --
        self.kampania_tabs = QTabWidget()
        self.kampania_tabs.addTab(self.search_tab, tr('🔍 Szukaj firm'))
        self.kampania_tabs.addTab(self.leads_tab, tr('📋 Lista leadów'))
        self.kampania_tabs.addTab(self.history_tab, tr('📜 Historia'))
        self.kampania_tabs.addTab(self.ai_tab, tr('🤖 Asystent AI'))

        # -- Grupa 2: WYSYŁKA (wysyłka + skrzynka odbiorcza) --
        self.wysylka_tabs = QTabWidget()
        self.wysylka_tabs.addTab(self.send_tab, tr('📤 Wyślij kampanię'))
        self.wysylka_tabs.addTab(self.inbox_tab, tr('📥 Skrzynka odbiorcza'))

        # -- Grupa 3: USTAWIENIA (settings_tab ma własne podzakładki) --

        self.tabs.addTab(self.kampania_tabs, tr('📁 Kampania'))
        self.tabs.addTab(self.wysylka_tabs, tr('📤 Wysyłka'))
        self.tabs.addTab(self.settings_tab, tr('⚙️ Ustawienia'))

        # AI workers
        self.ai_worker = None
        self.batch_ai_worker = None

        # Inicjalizuj AI providers
        ai_manager.register_provider("openai", OpenAIProvider())
        ai_manager.register_provider("gemini", GeminiProvider())
        ai_manager.register_provider("ollama", OllamaProvider())
        ai_manager.register_provider("lmstudio", LMStudioProvider())
        ai_manager.register_provider("deepseeklaude", DeepSeekLaudeProvider())

        # Budowanie zakładek
        self.build_search_tab()
        self.build_leads_tab()
        self.build_send_tab()
        self.build_settings_tab()
        self.build_history_tab()
        self.build_inbox_tab()
        self.build_ai_tab()

        # Podłącz sygnał zmiany profilu do odświeżenia UI
        self.profile_changed.connect(self._on_profile_changed)

        # Wczytaj dane dla bieżącego profilu
        self.load_settings()
        self._load_ai_config()
        self.refresh_leads()
        self.refresh_send_list()
        self.refresh_blacklist()
        self.refresh_profile_combo()

        # Wczytaj ostatni profil wewnętrzny (z bazy) – jeśli istnieje
        last_internal = db.get_setting("last_profile", "")
        if last_internal and last_internal in db.get_profile_names():
            self.load_profile(last_internal, silent=True)

    # ------------------------------------------------------------------
    # PAS NARZĘDZI Z PRZEŁĄCZNIKIEM PROFILI
    # ------------------------------------------------------------------
    def _create_toolbar(self):
        toolbar = QToolBar("Profile")
        toolbar.setMovable(False)
        self.addToolBar(toolbar)

        toolbar.addWidget(QLabel(tr('📁 Profil:')))

        self.profile_selector = QComboBox()
        self.profile_selector.setMinimumWidth(200)
        self.profile_selector.currentTextChanged.connect(self._on_profile_selector_changed)
        toolbar.addWidget(self.profile_selector)

        toolbar.addSeparator()

        # Przycisk odświeżenia listy profili
        btn_refresh_profiles = QPushButton(tr('🔄'))
        btn_refresh_profiles.setToolTip(tr('Odśwież listę profili'))
        btn_refresh_profiles.clicked.connect(self._refresh_profile_list)
        toolbar.addWidget(btn_refresh_profiles)

        toolbar.addSeparator()

        # Przycisk nowego profilu – uruchamia kreator
        btn_new_profile = QPushButton(tr('➕ Nowy'))
        btn_new_profile.setToolTip(tr('Utwórz nowy profil (kreator)'))
        btn_new_profile.clicked.connect(self._create_new_profile_wizard)
        toolbar.addWidget(btn_new_profile)

        # Przycisk kopiowania profilu
        btn_copy_profile = QPushButton(tr('📋 Kopiuj'))
        btn_copy_profile.setToolTip(tr('Kopiuj bieżący profil'))
        btn_copy_profile.clicked.connect(self._copy_current_profile_dialog)
        toolbar.addWidget(btn_copy_profile)

        # Przycisk usuwania profilu
        btn_delete_profile = QPushButton(tr('🗑 Usuń'))
        btn_delete_profile.setToolTip(tr('Usuń bieżący profil (nie można usunąć domyślnego)'))
        btn_delete_profile.clicked.connect(self._delete_current_profile)
        toolbar.addWidget(btn_delete_profile)

        toolbar.addSeparator()

        # Przycisk kreatora Szybki start
        btn_quickstart = QPushButton(tr('🚀 Szybki start'))
        btn_quickstart.setToolTip(tr('Przeprowadzi Cię krok po kroku przez konfigurację'))
        btn_quickstart.clicked.connect(self._run_quickstart_wizard)
        toolbar.addWidget(btn_quickstart)

        toolbar.addSeparator()

        # Etykieta informacyjna o ścieżce
        self.profile_path_label = QLabel("")
        self.profile_path_label.setStyleSheet("color: #888; font-size: 10px;")
        toolbar.addWidget(self.profile_path_label)

        # Wypełnij listę profili
        self._refresh_profile_list()

    def _refresh_profile_list(self):
        """Odświeża listę profili w comboboxie."""
        current = self.profile_selector.currentText()
        profiles = get_all_profiles()
        self.profile_selector.blockSignals(True)
        self.profile_selector.clear()
        self.profile_selector.addItems(profiles)
        # Ustaw bieżący profil
        active = get_current_profile_name()
        if active and active in profiles:
            self.profile_selector.setCurrentText(active)
        elif profiles:
            self.profile_selector.setCurrentIndex(0)
        self.profile_selector.blockSignals(False)
        self._update_profile_path_label()

    def _update_profile_path_label(self):
        """Aktualizuje etykietę z ścieżką profilu."""
        name = get_current_profile_name()
        if name:
            path = get_profile_path(name)
            self.profile_path_label.setText(f"📂 {path}")
        else:
            self.profile_path_label.setText("")

    def _on_profile_selector_changed(self, new_profile: str):
        """Obsługuje zmianę profilu z comboboxa."""
        if not new_profile:
            return
        current = get_current_profile_name()
        if current == new_profile:
            return
        # Przełącz
        if switch_profile(new_profile):
            self._current_profile = new_profile
            self.profile_changed.emit(new_profile)
            self._update_profile_path_label()
            # Odśwież tytuł okna
            self.setWindowTitle(f"PopeyLeadSonar – 🎯 {new_profile} – Hybrydowy System Leadów")
        else:
            QMessageBox.warning(self, tr('Błąd'), f"Nie udało się przełączyć na profil '{new_profile}'")
            # Przywróć poprzedni
            self._refresh_profile_list()

    def _on_profile_changed(self, profile_name: str):
        """Sygnał wywoływany po zmianie profilu – odświeża cały UI."""
        logger.info("Profil zmieniony na: %s – odświeżam UI", profile_name)
        # Zamknij aktywne wątki (jeśli działają)
        if self.search_worker:
            self.search_worker.stop()
        if self.send_worker:
            self.send_worker.stop()
        if self.autopilot_worker:
            self.autopilot_worker.stop()
        # Odśwież wszystkie widoki
        self.load_settings()
        self._load_ai_config()
        self._refresh_inbox_accounts()
        self.refresh_leads()
        self.refresh_send_list()
        self.refresh_blacklist()
        self.refresh_history()
        # Odśwież listę profili wewnętrznych (z bazy)
        self.refresh_profile_combo()
        # Wczytaj ostatni profil wewnętrzny
        last_internal = db.get_setting("last_profile", "")
        if last_internal and last_internal in db.get_profile_names():
            self.load_profile(last_internal, silent=True)
        # Odśwież statystyki (jeśli wykres)
        if MATPLOTLIB_AVAILABLE and hasattr(self, 'stats_canvas') and self.stats_canvas:
            self.refresh_stats()

    # ------------------------------------------------------------------
    # Dialogi zarządzania profilami
    # ------------------------------------------------------------------
    def _run_quickstart_wizard(self):
        """Uruchamia kreator Szybki start i odświeża UI zapisanymi danymi."""
        wizard = QuickStartWizard(self)
        if wizard.exec() == QWizard.Accepted:
            self.load_settings()
            self._load_ai_config()
            settings = get_current_profile_settings()
            self.queries_edit.setPlainText(settings.get("last_queries", DEFAULT_QUERIES))
            self.locations_edit.setPlainText(settings.get("last_locations", DEFAULT_LOCATIONS))
            QMessageBox.information(
                self, tr('Gotowe'),
                tr('Ustawienia zapisane! Przejdź do zakładki 🔍 Wyszukiwanie, żeby znaleźć pierwsze firmy.')
            )

    def _create_new_profile_wizard(self):
        """Uruchamia kreator nowego profilu i odświeża UI po sukcesie."""
        wizard = ProfileWizard(self)
        if wizard.exec() == QWizard.Accepted:
            # Kreator zakończony sukcesem – odśwież UI
            self._refresh_profile_list()
            current = get_current_profile_name()
            if current:
                self.profile_selector.setCurrentText(current)
                self.profile_changed.emit(current)
                self._on_profile_changed(current)
            else:
                self._refresh_profile_list()

    def _copy_current_profile_dialog(self):
        current = get_current_profile_name()
        if not current:
            return
        new_name, ok = QInputDialog.getText(
            self, tr('Kopiuj profil'),
            f"Kopiuj profil '{current}' pod nazwą:"
        )
        if not ok or not new_name.strip():
            return
        new_name = new_name.strip()
        if new_name in get_all_profiles():
            QMessageBox.warning(self, tr('Błąd'), f"Profil '{new_name}' już istnieje.")
            return
        if copy_profile(current, new_name):
            self._refresh_profile_list()
            self.profile_selector.setCurrentText(new_name)
            QMessageBox.information(self, tr('OK'), f"Profil '{current}' skopiowano jako '{new_name}'.")
        else:
            QMessageBox.warning(self, tr('Błąd'), tr('Nie udało się skopiować profilu.'))

    def _delete_current_profile(self):
        current = get_current_profile_name()
        if not current:
            return
        if current == DEFAULT_PROFILE_NAME:
            QMessageBox.warning(self, tr('Błąd'), tr('Nie można usunąć domyślnego profilu.'))
            return
        reply = QMessageBox.question(
            self, tr('Potwierdzenie'),
            f"Czy na pewno usunąć profil '{current}'?\nWszystkie dane (leady, historię, ustawienia) zostaną trwale usunięte.",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.No:
            return
        if delete_profile_by_name(current):
            self._refresh_profile_list()
            # Po usunięciu przełączono na domyślny
            self._current_profile = get_current_profile_name()
            self.profile_changed.emit(self._current_profile)
            QMessageBox.information(self, tr('OK'), f"Profil '{current}' usunięty.")
        else:
            QMessageBox.warning(self, tr('Błąd'), tr('Nie udało się usunąć profilu.'))

    # ------------------------------------------------------------------
    # ZAKŁADKA: SZUKAJ
    # ------------------------------------------------------------------
    def build_search_tab(self):
        container = QWidget()
        layout = QVBoxLayout(container)

        # Profil wewnętrzny (kampania)
        profile_group = QGroupBox(tr('👤 Profil kampanii (kategorie + lokalizacje + szablon + temat)'))
        profile_layout = QHBoxLayout(profile_group)
        self.profile_combo = QComboBox()
        self.profile_combo.setMinimumWidth(200)
        profile_layout.addWidget(self.profile_combo, stretch=1)

        btn_load_profile = QPushButton(tr('📂 Wczytaj'))
        btn_load_profile.clicked.connect(lambda: self.load_profile(self.profile_combo.currentText()))
        profile_layout.addWidget(btn_load_profile)

        btn_save_profile = QPushButton(tr('💾 Nadpisz'))
        btn_save_profile.clicked.connect(self.overwrite_current_profile)
        profile_layout.addWidget(btn_save_profile)

        btn_save_as_profile = QPushButton(tr('💾 Zapisz jako nowy...'))
        btn_save_as_profile.clicked.connect(self.save_as_new_profile)
        profile_layout.addWidget(btn_save_as_profile)

        btn_delete_profile = QPushButton(tr('🗑 Usuń'))
        btn_delete_profile.clicked.connect(self.delete_current_internal_profile)
        profile_layout.addWidget(btn_delete_profile)
        layout.addWidget(profile_group)

        layout.addWidget(QLabel(tr('📌 Kategorie (jedna na linię):')))
        self.queries_edit = QTextEdit()
        self.queries_edit.setMaximumHeight(200)
        self.queries_edit.setPlainText(DEFAULT_QUERIES)
        self.queries_edit.textChanged.connect(self._auto_save_search_params)
        layout.addWidget(self.queries_edit)

        layout.addWidget(QLabel(tr('📍 Lokalizacje (jedna na linię):')))
        self.locations_edit = QTextEdit()
        self.locations_edit.setMaximumHeight(200)
        self.locations_edit.setPlainText(DEFAULT_LOCATIONS)
        self.locations_edit.textChanged.connect(self._auto_save_search_params)
        layout.addWidget(self.locations_edit)

        row = QHBoxLayout()
        row.addWidget(QLabel(tr('Limit na zapytanie:')))
        self.limit_spin = QSpinBox()
        self.limit_spin.setRange(3, 50)
        self.limit_spin.setValue(15)
        row.addWidget(self.limit_spin)

        row.addWidget(QLabel(tr('Czas szukania (min, 0=bez limitu):')))
        self.czas_spin = QSpinBox()
        self.czas_spin.setRange(0, 120)
        self.czas_spin.setValue(60)
        row.addWidget(self.czas_spin)
        layout.addLayout(row)

        self.force_research_check = QCheckBox(
            tr('🔁 Szukaj od nowa (również w kategoriach/lokalizacjach już wcześniej przeszukanych)')
        )
        self.force_research_check.setToolTip(
            tr('Domyślnie program pomija kombinacje kategoria+lokalizacja, które już kiedyś przeszukał w tym profilu, żeby nie tracić czasu na to samo od początku. Zaznacz to, żeby wymusić pełne przeszukanie wszystkiego od nowa.')
        )
        layout.addWidget(self.force_research_check)

        row2 = QHBoxLayout()
        self.search_btn = QPushButton(tr('🔍 ROZPOCZNIJ SZUKANIE'))
        self.search_btn.setStyleSheet("background: #2b5e2b; font-weight: bold; font-size: 14px; padding: 10px;")
        self.search_btn.clicked.connect(self.start_search)
        row2.addWidget(self.search_btn)

        self.stop_search_btn = QPushButton(tr('⏹️ STOP'))
        self.stop_search_btn.setStyleSheet("background: #8b0000; font-weight: bold;")
        self.stop_search_btn.clicked.connect(self.stop_search)
        self.stop_search_btn.setEnabled(False)
        row2.addWidget(self.stop_search_btn)

        self.auto_save_check = QCheckBox(tr('💾 Automatycznie zapisuj do bazy'))
        self.auto_save_check.setChecked(True)
        row2.addWidget(self.auto_save_check)

        # Lead Scoring during search
        self.ai_score_during_search = QCheckBox(tr('🤖 AI Lead Scoring'))
        self.ai_score_during_search.setToolTip(tr('Oceniaj leady podczas szukania za pomocą AI'))
        row2.addWidget(self.ai_score_during_search)
        layout.addLayout(row2)

        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        self.search_status = QLabel(tr('Gotowy'))
        layout.addWidget(self.search_status)

        layout.addWidget(QLabel(tr('📋 Znalezione firmy (z emailami):')))
        self.search_results = QListWidget()
        layout.addWidget(self.search_results)

        self.counter_label = QLabel(tr('Znaleziono: 0 firm'))
        layout.addWidget(self.counter_label)

        btn_save = QPushButton(tr('💾 Zapisz wszystkie do bazy (ręcznie)'))
        btn_save.clicked.connect(self.save_all_found)
        layout.addWidget(btn_save)

        # Tryb automatyczny
        auto_group = QGroupBox(tr('🤖 Tryb automatyczny (szuka i od razu wysyła)'))
        auto_layout = QVBoxLayout(auto_group)
        auto_hint = QLabel(
            tr('Używa szablonu/tematu z zakładki 📤 Wysyłka i konta SMTP z ⚙️ Ustawienia.')
        )
        auto_hint.setWordWrap(True)
        auto_hint.setStyleSheet("color: #a6adc8; font-size: 11px;")
        auto_layout.addWidget(auto_hint)

        auto_btn_row = QHBoxLayout()
        self.autopilot_btn = QPushButton(tr('🤖 START (szukaj + wysyłaj automatycznie)'))
        self.autopilot_btn.setStyleSheet("background: #5e2b8b; font-weight: bold; padding: 8px;")
        self.autopilot_btn.clicked.connect(self.start_autopilot)
        auto_btn_row.addWidget(self.autopilot_btn)

        self.autopilot_stop_btn = QPushButton(tr('⏹️ STOP'))
        self.autopilot_stop_btn.setStyleSheet("background: #8b0000; font-weight: bold;")
        self.autopilot_stop_btn.clicked.connect(self.stop_autopilot)
        self.autopilot_stop_btn.setEnabled(False)
        auto_btn_row.addWidget(self.autopilot_stop_btn)
        auto_layout.addLayout(auto_btn_row)

        self.autopilot_status = QLabel(tr('Gotowy'))
        auto_layout.addWidget(self.autopilot_status)
        self.autopilot_counters = QLabel(tr('Znaleziono: 0 | Wysłano: 0 | Błędów: 0'))
        auto_layout.addWidget(self.autopilot_counters)
        layout.addWidget(auto_group)

        # Scroll
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(container)
        self.search_tab.layout().addWidget(scroll)

    # ------------------------------------------------------------------
    # METODY POMOCNICZE DLA PROFILI WEWNĘTRZNYCH
    # ------------------------------------------------------------------
    def refresh_profile_combo(self, select: Optional[str] = None):
        names = db.get_profile_names()
        self.profile_combo.blockSignals(True)
        self.profile_combo.clear()
        self.profile_combo.addItems(names)
        if select and select in names:
            self.profile_combo.setCurrentText(select)
        self.profile_combo.blockSignals(False)

    def load_profile(self, name: str, silent: bool = False):
        profile = db.get_profile(name)
        if not profile:
            if not silent:
                QMessageBox.warning(self, tr('Błąd'), f"Nie znaleziono profilu '{name}'.")
            return
        # Jeśli profil wewnętrzny ma puste pola, użyj ustawień z profilu (last_template, etc.)
        settings = get_current_profile_settings()
        queries = profile["queries"] or settings.get("last_queries", DEFAULT_QUERIES)
        locations = profile["locations"] or settings.get("last_locations", DEFAULT_LOCATIONS)
        template = profile["template"] or settings.get("last_template", DEFAULT_TEMPLATE)
        subject = profile["subject"] or settings.get("last_subject", DEFAULT_SUBJECT)
        self.queries_edit.setPlainText(queries)
        self.locations_edit.setPlainText(locations)
        self.szablon_edit.setPlainText(template)
        self.temat_edit.setText(subject)
        self.refresh_profile_combo(select=name)
        db.set_setting("last_profile", name)
        if not silent:
            self.search_status.setText(f"📂 Wczytano profil kampanii: {name}")

    def _current_internal_profile_fields(self):
        return (
            self.queries_edit.toPlainText(),
            self.locations_edit.toPlainText(),
            self.szablon_edit.toPlainText(),
            self.temat_edit.text(),
        )

    def overwrite_current_profile(self):
        name = self.profile_combo.currentText()
        if not name:
            QMessageBox.warning(self, tr('Błąd'), tr('Brak wybranego profilu kampanii.'))
            return
        reply = QMessageBox.question(
            self, tr('Potwierdzenie'), f"Nadpisać profil kampanii '{name}'?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.No:
            return
        queries, locations, template, subject = self._current_internal_profile_fields()
        db.save_profile(name, queries, locations, template, subject)
        db.set_setting("last_profile", name)
        # Zapisz też do profilu (dla szybszego wczytania)
        settings = get_current_profile_settings()
        settings["last_queries"] = queries
        settings["last_locations"] = locations
        settings["last_template"] = template
        settings["last_subject"] = subject
        update_current_profile_settings(settings)
        QMessageBox.information(self, tr('OK'), f"Profil kampanii '{name}' zaktualizowany.")

    def save_as_new_profile(self):
        name, ok = QInputDialog.getText(self, tr('Nowy profil kampanii'), tr('Nazwa nowego profilu:'))
        name = name.strip()
        if not ok or not name:
            return
        if name in db.get_profile_names():
            QMessageBox.warning(self, tr('Błąd'), f"Profil '{name}' już istnieje.")
            return
        queries, locations, template, subject = self._current_internal_profile_fields()
        db.save_profile(name, queries, locations, template, subject)
        db.set_setting("last_profile", name)
        self.refresh_profile_combo(select=name)
        QMessageBox.information(self, tr('OK'), f"Zapisano nowy profil kampanii: {name}")

    def delete_current_internal_profile(self):
        name = self.profile_combo.currentText()
        if not name:
            return
        names = db.get_profile_names()
        if len(names) <= 1:
            QMessageBox.warning(self, tr('Błąd'), tr('Nie można usunąć ostatniego profilu kampanii.'))
            return
        reply = QMessageBox.question(
            self, tr('Potwierdzenie'), f"Usunąć profil kampanii '{name}'?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.No:
            return
        db.delete_profile_from_db(name)
        self.refresh_profile_combo()
        remaining = db.get_profile_names()
        if remaining:
            self.load_profile(remaining[0], silent=True)

    # ------------------------------------------------------------------
    # WYSZUKIWANIE (start_search, stop_search, itp.)
    # ------------------------------------------------------------------
    def start_search(self):
        queries = [q.strip() for q in self.queries_edit.toPlainText().strip().split('\n') if q.strip()]
        locations = [l.strip() for l in self.locations_edit.toPlainText().strip().split('\n') if l.strip()]

        if not queries or not locations:
            QMessageBox.warning(self, tr('Błąd'), tr('Wpisz kategorie i lokalizacje!'))
            return

        self.search_btn.setEnabled(False)
        self.stop_search_btn.setEnabled(True)
        self.progress_bar.setRange(0, 100)
        self.search_results.clear()
        self.found_leads = []
        self.counter_label.setText(tr('Znaleziono: 0 firm'))
        self.search_status.setText(tr('⏳ Szukam...'))

        proxy_enabled = get_current_profile_settings().get("proxy_enabled", False)
        raw_proxies = get_current_profile_settings().get("proxy_list", "")
        proxies = [p.strip() for p in raw_proxies.splitlines() if p.strip()] if proxy_enabled else []
        ai_scoring_enabled = get_current_profile_settings().get("ai_scoring_enabled", False)

        self.search_worker = SearchWorker(
            queries, locations, self.limit_spin.value(), self.czas_spin.value(), proxies,
            force_research=self.force_research_check.isChecked(),
            ai_scoring=ai_scoring_enabled
        )
        self.search_worker.progress.connect(self.progress_bar.setValue)
        self.search_worker.status.connect(self.search_status.setText)
        self.search_worker.result.connect(self.on_search_result)
        self.search_worker.finished.connect(self.on_search_finished)
        self.search_worker.error.connect(lambda e: self.search_status.setText(f"❌ {e}"))
        self.search_worker.live_result.connect(self.on_live_result)
        self.search_worker.start()

    def stop_search(self):
        if self.search_worker:
            self.search_worker.stop()
            self.search_status.setText(tr('⏹️ Zatrzymywanie...'))

    def on_live_result(self, lead):
        email = lead.get('email', '')
        if not email:
            return
        if any(item.get('email') == email for item in self.found_leads):
            return

        self.found_leads.append(lead)
        self.counter_label.setText(f"Znaleziono: {len(self.found_leads)} firm")

        text = f"{lead.get('name', '')} | {email} | {lead.get('website', '')}"
        item = QListWidgetItem(text)
        item.setData(Qt.UserRole, lead)

        if email in db.get_wyslano_emails():
            item.setBackground(QColor(*COLOR_SENT_LIST_ALT))

        self.search_results.addItem(item)

        if self.auto_save_check.isChecked():
            db.add_lead(
                lead.get('name', ''), email, lead.get('address', ''),
                '', lead.get('website', ''), lead.get('category', '')
            )

    def on_search_result(self, results):
        self.search_status.setText(f"✅ Znaleziono łącznie {len(results)} firm z emailami")

    def on_search_finished(self):
        self.search_btn.setEnabled(True)
        self.stop_search_btn.setEnabled(False)
        self.search_status.setText(f"✅ Gotowe! Znaleziono {len(self.found_leads)} firm")
        self.refresh_leads()
        self.refresh_send_list()

    def save_all_found(self):
        if not self.found_leads:
            QMessageBox.warning(self, tr('Błąd'), tr('Brak leadów do zapisania!'))
            return

        saved = sum(
            1 for lead in self.found_leads
            if lead.get('email') and db.add_lead(
                lead.get('name', ''), lead['email'], lead.get('address', ''),
                '', lead.get('website', ''), lead.get('category', '')
            )
        )

        self.refresh_leads()
        self.refresh_send_list()
        QMessageBox.information(self, tr('OK'), f"Zapisano {saved} nowych leadów")

    # ------------------------------------------------------------------
    # TRYB AUTOMATYCZNY
    # ------------------------------------------------------------------
    def start_autopilot(self):
        queries = [q.strip() for q in self.queries_edit.toPlainText().strip().split('\n') if q.strip()]
        locations = [l.strip() for l in self.locations_edit.toPlainText().strip().split('\n') if l.strip()]
        if not queries or not locations:
            QMessageBox.warning(self, tr('Błąd'), tr('Wpisz kategorie i lokalizacje!'))
            return

        szablon = self.szablon_edit.toPlainText().strip()
        temat = self.temat_edit.text().strip()
        if not szablon or not temat:
            QMessageBox.warning(self, tr('Błąd'), tr('Wpisz szablon i temat w zakładce 📤 Wysyłka!'))
            return

        # Pobierz ustawienia z profilu
        settings = get_current_profile_settings()
        gmail_user = settings.get("gmail_user", "")
        gmail_pass = settings.get("gmail_password", "")
        if not gmail_user or not gmail_pass:
            QMessageBox.warning(self, tr('Błąd'), tr('Skonfiguruj Gmail/SMTP w ustawieniach profilu!'))
            return

        smtp_host = settings.get("smtp_host", SMTP_RELAY_HOST)
        smtp_port = settings.get("smtp_port", SMTP_RELAY_PORT)

        ok, msg = test_gmail_connection(gmail_user, gmail_pass, smtp_host, smtp_port)
        if not ok:
            QMessageBox.warning(self, tr('Błąd SMTP'), f"Nie można połączyć: {msg}")
            return

        session_cap = settings.get("dzienny_limit", SESSION_HARD_CAP)
        custom_delay = settings.get("custom_send_delay", CUSTOM_SEND_DELAY_DEFAULT)
        custom_cap = settings.get("custom_session_cap", CUSTOM_SESSION_CAP_DEFAULT)

        proxy_enabled = settings.get("proxy_enabled", False)
        raw_proxies = settings.get("proxy_list", "")
        proxies = [p.strip() for p in raw_proxies.splitlines() if p.strip()] if proxy_enabled else []

        send_delay = get_send_delay(smtp_host, custom_delay)
        reply = QMessageBox.question(
            self, tr('Potwierdzenie'),
            f"Tryb automatyczny będzie SZUKAĆ nowych leadów i OD RAZU wysyłać do nich "
            f"wiadomość (tempo: {send_delay:g} s/wiadomość). Kontynuować?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.No:
            return

        html = settings.get("html_enabled", False)
        verify_mx = settings.get("mx_verify_enabled", False)
        smime_sign = settings.get("smime_enabled", False)
        attachments = [p.strip() for p in settings.get("attachments", "").split(',') if p.strip()]
        use_rotation = settings.get("account_rotation_enabled", False)
        rotator = None
        if use_rotation:
            accs = db.get_smtp_accounts()
            if accs:
                rotator = SMTPAccountRotator(accs)
                rotator.set_max_per_account(settings.get("rotation_max_per_account", 1000))

        self.autopilot_worker = AutoPilotWorker(
            queries, locations, self.limit_spin.value(), szablon, temat,
            gmail_user, gmail_pass, smtp_host, smtp_port, session_cap, proxies,
            custom_delay, custom_cap,
            html=html, attachments=attachments,
            verify_mx=verify_mx, check_blacklist=True,
            smime_sign=smime_sign,
            personalized_attachments={},
            use_account_rotation=use_rotation,
            rotator=rotator,
            dry_run=self.dry_run_check.isChecked()
        )
        self.autopilot_worker.status.connect(self.autopilot_status.setText)
        self.autopilot_worker.counters.connect(self.on_autopilot_counters)
        self.autopilot_worker.lead_processed.connect(lambda _l: (self.refresh_leads(), self.refresh_send_list()))
        self.autopilot_worker.error.connect(lambda e: self.autopilot_status.setText(f"❌ {e}"))
        self.autopilot_worker.finished.connect(self.on_autopilot_finished)
        self.autopilot_worker.start()

        self.autopilot_btn.setEnabled(False)
        self.autopilot_stop_btn.setEnabled(True)
        self.search_btn.setEnabled(False)
        self.autopilot_status.setText(tr('⏳ Tryb automatyczny wystartował...'))

    def stop_autopilot(self):
        if self.autopilot_worker:
            self.autopilot_worker.stop()
            self.autopilot_status.setText(tr('⏹️ Zatrzymywanie...'))

    def on_autopilot_counters(self, found, sent, errors):
        self.autopilot_counters.setText(f"Znaleziono: {found} | Wysłano: {sent} | Błędów: {errors}")

    def on_autopilot_finished(self):
        self.autopilot_btn.setEnabled(True)
        self.autopilot_stop_btn.setEnabled(False)
        self.search_btn.setEnabled(True)
        self.refresh_leads()
        self.refresh_send_list()

    # ------------------------------------------------------------------
    # ZAKŁADKA: LEADZY
    # ------------------------------------------------------------------
    def build_leads_tab(self):
        container = QWidget()
        layout = QVBoxLayout(container)

        row = QHBoxLayout()
        self.filter_status = QComboBox()
        self.filter_status.addItems(["Wszystkie", "nowy", "wysłano", "błąd"])
        self.filter_status.currentTextChanged.connect(self.refresh_leads)
        row.addWidget(QLabel(tr('Status:')))
        row.addWidget(self.filter_status)

        self.filter_search = QLineEdit()
        self.filter_search.setPlaceholderText(tr('Szukaj...'))
        self.filter_search.textChanged.connect(self.refresh_leads)
        row.addWidget(self.filter_search)

        btn_refresh = QPushButton(tr('🔄 Odśwież'))
        btn_refresh.clicked.connect(self.refresh_leads)
        row.addWidget(btn_refresh)

        btn_delete_sent = QPushButton(tr('🗑 Usuń wszystkie wysłane'))
        btn_delete_sent.clicked.connect(self.delete_sent)
        row.addWidget(btn_delete_sent)

        btn_blacklist = QPushButton(tr('🚫 Dodaj zaznaczone do blacklist'))
        btn_blacklist.clicked.connect(self.add_selected_to_blacklist)
        row.addWidget(btn_blacklist)

        layout.addLayout(row)

        self.leads_table = QTableWidget()
        self.leads_table.setColumnCount(7)
        self.leads_table.setHorizontalHeaderLabels(
            [tr('ID'), tr('Firma'), tr('Email'), tr('Adres'), tr('Typ'), tr('Status'), tr('Wysłano')]
        )
        self.leads_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.leads_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.leads_table)

        row2 = QHBoxLayout()
        btn_export = QPushButton(tr('📤 Eksportuj CSV'))
        btn_export.clicked.connect(self.export_csv)
        row2.addWidget(btn_export)

        btn_import = QPushButton(tr('📥 Importuj CSV/Excel'))
        btn_import.clicked.connect(self.import_csv_excel)
        row2.addWidget(btn_import)

        layout.addLayout(row2)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(container)
        self.leads_tab.layout().addWidget(scroll)

    def refresh_leads(self):
        status = self.filter_status.currentText()
        status = None if status == "Wszystkie" else status
        search = self.filter_search.text().strip().lower()
        rows = db.get_leads_summary(status=status, search=search)

        self.leads_table.setRowCount(len(rows))
        STATUS_COL = 5
        for i, row in enumerate(rows):
            for col, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val else "")
                if col == STATUS_COL:
                    if val == 'wysłano':
                        item.setBackground(QColor(*COLOR_OK))
                    elif val == 'błąd':
                        item.setBackground(QColor(*COLOR_ERROR))
                self.leads_table.setItem(i, col, item)
        self.leads_table.resizeColumnsToContents()

    def delete_sent(self):
        reply = QMessageBox.question(
            self, tr('Potwierdzenie'), tr('Usunąć wszystkie wysłane leady?'),
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            db.delete_sent_leads()
            self.refresh_leads()
            self.refresh_send_list()

    def export_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, "Zapisz CSV", "leady.csv", "CSV Files (*.csv)")
        if not path:
            return
        rows = db.get_leads()
        headers = ["id", "firma", "kontakt", "email", "adres", "telefon", "website", "typ", "status", "wyslano"]
        try:
            with open(path, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(rows)
            QMessageBox.information(self, tr('OK'), f"Eksportowano {len(rows)} leadów")
        except OSError as e:
            QMessageBox.critical(self, tr('Błąd'), f"Nie można zapisać pliku: {e}")

    def import_csv_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Wybierz plik", "",
            "Pliki CSV/Excel (*.csv *.xlsx *.xls)"
        )
        if not path:
            return
        try:
            imported = 0
            if path.endswith('.csv'):
                with open(path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        email = (row.get('email') or '').strip()
                        if email:
                            if db.add_lead(
                                row.get('firma', ''), email, row.get('adres', ''),
                                row.get('telefon', ''), row.get('website', ''),
                                row.get('typ', ''), row.get('kontakt', '')
                            ):
                                imported += 1
            else:
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(path, read_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    col_map = {h.lower(): i for i, h in enumerate(headers)}
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        email = row[col_map.get('email', -1)] if 'email' in col_map else None
                        if email:
                            firma = row[col_map.get('firma', -1)] if 'firma' in col_map else ''
                            adres = row[col_map.get('adres', -1)] if 'adres' in col_map else ''
                            telefon = row[col_map.get('telefon', -1)] if 'telefon' in col_map else ''
                            website = row[col_map.get('website', -1)] if 'website' in col_map else ''
                            typ = row[col_map.get('typ', -1)] if 'typ' in col_map else ''
                            kontakt = row[col_map.get('kontakt', -1)] if 'kontakt' in col_map else ''
                            if db.add_lead(firma, email, adres, telefon, website, typ, kontakt):
                                imported += 1
                except ImportError:
                    QMessageBox.warning(self, tr('Błąd'), tr('Zainstaluj openpyxl: pip install openpyxl'))
                    return

            self.refresh_leads()
            self.refresh_send_list()
            QMessageBox.information(self, tr('OK'), f"Zaimportowano {imported} leadów")
        except Exception as e:
            QMessageBox.critical(self, tr('Błąd'), f"Nie można zaimportować: {e}")

    def add_selected_to_blacklist(self):
        selected_rows = set()
        for item in self.leads_table.selectedItems():
            selected_rows.add(item.row())
        count = 0
        for row in selected_rows:
            email_item = self.leads_table.item(row, 2)
            if email_item and email_item.text():
                if db.add_to_blacklist(email_item.text(), "manual"):
                    count += 1
        self.refresh_blacklist()
        QMessageBox.information(self, tr('OK'), f"Dodano {count} adresów do blacklist.")

    # ------------------------------------------------------------------
    # ZAKŁADKA: WYSYŁKA
    # ------------------------------------------------------------------
    def build_send_tab(self):
        container = QWidget()
        layout = QVBoxLayout(container)

        dry_run_box = QGroupBox(tr("🎬 Tryb demo / nagrywanie"))
        dry_run_layout = QVBoxLayout(dry_run_box)
        self.dry_run_check = QCheckBox(
            tr("Dry-run: NIE wysyłaj prawdziwych maili (symuluj całą wysyłkę - AI, "
               "opóźnienia, statusy - ale bez łączenia się z SMTP)")
        )
        self.dry_run_check.setStyleSheet("font-weight: bold;")
        self.dry_run_check.setToolTip(
            tr("Włącz przed nagrywaniem filmiku demo. Wszystko dzieje się normalnie "
               "(AI pisze maile, UI się aktualizuje, statusy się zmieniają), ale żaden "
               "e-mail faktycznie nie zostanie wysłany - nie trzeba nawet podawać "
               "prawdziwego hasła SMTP. Dotyczy zwykłej wysyłki, AI Auto Send i Autopilota.")
        )
        dry_run_layout.addWidget(self.dry_run_check)
        layout.addWidget(dry_run_box)

        layout.addWidget(QLabel(tr('📝 Szablon wiadomości:')))
        self.szablon_edit = QTextEdit()
        self.szablon_edit.setMaximumHeight(200)
        self.szablon_edit.setPlainText(DEFAULT_TEMPLATE)
        self.szablon_edit.textChanged.connect(self._auto_save_template_subject)
        layout.addWidget(self.szablon_edit)

        template_save_row = QHBoxLayout()
        btn_save_template = QPushButton(tr("💾 Zapisz szablon"))
        btn_save_template.setToolTip(
            tr("Zapisuje aktualną treść szablonu i temat na stałe (per profil), żeby "
               "przy kolejnym uruchomieniu programu nie trzeba było wpisywać ich od nowa.")
        )
        btn_save_template.clicked.connect(self._manual_save_template)
        template_save_row.addWidget(btn_save_template)
        template_save_row.addStretch()
        layout.addLayout(template_save_row)

        hint = QLabel(
            tr('ℹ️ Zmienne: {firma}, {kontakt}, {email}, {adres}, {telefon}, {id}\nMieszacz treści: {{wariant 1|wariant 2|wariant 3}}')
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color: #a6adc8; font-size: 11px;")
        layout.addWidget(hint)

        layout.addWidget(QLabel(tr('📌 Temat:')))
        self.temat_edit = QLineEdit()
        self.temat_edit.setPlaceholderText(DEFAULT_SUBJECT)
        self.temat_edit.setText(DEFAULT_SUBJECT)
        self.temat_edit.textChanged.connect(self._auto_save_template_subject)
        layout.addWidget(self.temat_edit)

        # Opcje wysyłki + Lead Score Filter
        options_group = QGroupBox(tr('⚙️ Opcje wysyłki'))
        options_layout = QVBoxLayout(options_group)

        options_row = QHBoxLayout()
        self.html_check = QCheckBox(tr('HTML'))
        options_row.addWidget(self.html_check)
        self.mx_check = QCheckBox(tr('Sprawdzaj MX'))
        options_row.addWidget(self.mx_check)
        self.smime_check = QCheckBox(tr('S/MIME'))
        options_row.addWidget(self.smime_check)

        # Lead Score filtering
        options_row.addWidget(QLabel(tr('Min Lead Score:')))
        self.lead_score_min_spin = QSpinBox()
        self.lead_score_min_spin.setRange(-1, 100)
        self.lead_score_min_spin.setValue(0)
        self.lead_score_min_spin.setSuffix(tr('%'))
        self.lead_score_min_spin.setToolTip(
            tr('Leady z oceną AI poniżej tego progu nie zostaną wysłane. Ta sama wartość co „Próg punktacji minimalnej” w Ustawieniach -> AI - zmiana tutaj zapisuje się tam automatycznie i na odwrót.')
        )
        self.lead_score_min_spin.valueChanged.connect(self._auto_save_lead_score_threshold)
        options_row.addWidget(self.lead_score_min_spin)

        self.lead_score_allow_unscored = QCheckBox(tr('Wysyłaj też nieocenione (score = -1)'))
        self.lead_score_allow_unscored.setChecked(True)
        self.lead_score_allow_unscored.setToolTip(
            tr('Jeśli odznaczone: leady, których AI jeszcze nie oceniło (albo ocena się nie powiodła), NIE zostaną wysłane, dopóki nie dostaną realnego wyniku.')
        )
        self.lead_score_allow_unscored.toggled.connect(self._auto_save_lead_score_threshold)
        options_row.addWidget(self.lead_score_allow_unscored)

        options_layout.addLayout(options_row)
        layout.addWidget(options_group)

        # Załączniki
        attach_group = QGroupBox(tr('📎 Załączniki (oddzielone przecinkami)'))
        attach_layout = QVBoxLayout(attach_group)
        self.attachments_edit = QLineEdit()
        self.attachments_edit.setPlaceholderText(tr('ścieżka/do/pliku1.pdf, ścieżka/do/pliku2.jpg'))
        attach_layout.addWidget(self.attachments_edit)
        layout.addWidget(attach_group)

        btn_preview = QPushButton(tr('👁 Podgląd wiadomości'))
        btn_preview.clicked.connect(self.preview_template)
        layout.addWidget(btn_preview)

        layout.addWidget(QLabel(tr('📋 Wybierz leady do wysyłki:')))
        self.send_list = QListWidget()
        self.send_list.setSelectionMode(QListWidget.MultiSelection)
        layout.addWidget(self.send_list)

        row = QHBoxLayout()
        btn_select_all = QPushButton(tr('✅ Zaznacz wszystkie'))
        btn_select_all.clicked.connect(lambda: self.send_list.selectAll())
        row.addWidget(btn_select_all)
        btn_select_new = QPushButton(tr('🆕 Tylko nowe'))
        btn_select_new.clicked.connect(self.select_new_only)
        row.addWidget(btn_select_new)
        btn_select_none = QPushButton(tr('❌ Odznacz'))
        btn_select_none.clicked.connect(lambda: self.send_list.clearSelection())
        row.addWidget(btn_select_none)
        layout.addLayout(row)

        row2 = QHBoxLayout()
        btn_send_100 = QPushButton(tr('📤 Wyślij 100'))
        btn_send_100.clicked.connect(lambda: self.start_send(100))
        btn_send_100.setStyleSheet("background: #2b5e2b; font-weight: bold;")
        row2.addWidget(btn_send_100)
        btn_send_200 = QPushButton(tr('📤 Wyślij 200'))
        btn_send_200.clicked.connect(lambda: self.start_send(200))
        btn_send_200.setStyleSheet("background: #2b5e2b; font-weight: bold;")
        row2.addWidget(btn_send_200)
        btn_send_all = QPushButton(tr('📤 Wyślij WSZYSTKIE'))
        btn_send_all.clicked.connect(lambda: self.start_send(9999))
        btn_send_all.setStyleSheet("background: #8b5e2b; font-weight: bold;")
        row2.addWidget(btn_send_all)
        self.stop_send_btn = QPushButton(tr('⏹️ STOP'))
        self.stop_send_btn.clicked.connect(self.stop_send)
        self.stop_send_btn.setEnabled(False)
        self.stop_send_btn.setStyleSheet("background: #8b0000; font-weight: bold;")
        row2.addWidget(self.stop_send_btn)
        layout.addLayout(row2)

        # W pelni automatyczna wysylka AI: czyta strone kazdego leada, ocenia
        # czy moze potrzebowac naszej uslugi, pisze spersonalizowany mail i
        # wysyla - bez zadnego potwierdzania per-lead.
        ai_auto_group = QGroupBox(tr('🤖 AI Auto-Send (czyta stronę + ocenia + pisze + wysyła - w pełni automatycznie)'))
        ai_auto_layout = QVBoxLayout(ai_auto_group)
        ai_auto_hint = QLabel(
            tr('Dla każdego zaznaczonego leada: pobiera treść jego strony WWW, ocenia AI czy firma może potrzebować Waszej usługi (opisanej w ⚙️ Ustawienia -> Dane firmy -> "Czym się zajmujemy"), i jeśli tak - pisze spersonalizowany mail i wysyła. Stopka z danymi firmy i telefonem dodawana jest automatycznie, niezależnie od AI. Leady poniżej progu NIE są wysyłane.')
        )
        ai_auto_hint.setWordWrap(True)
        ai_auto_hint.setStyleSheet("color: #a6adc8; font-size: 11px;")
        ai_auto_layout.addWidget(ai_auto_hint)

        ai_auto_row = QHBoxLayout()
        ai_auto_row.addWidget(QLabel(tr('Branża odbiorców:')))
        self.ai_auto_industry_edit = QLineEdit()
        self.ai_auto_industry_edit.setPlaceholderText(tr('np. gastronomia, IT, budownictwo'))
        ai_auto_row.addWidget(self.ai_auto_industry_edit)
        ai_auto_row.addWidget(QLabel(tr("Język maili:")))
        self.ai_auto_language_combo = QComboBox()
        self.ai_auto_language_combo.addItem(tr("Auto (dopasuj do firmy)"), "auto")
        self.ai_auto_language_combo.addItem(tr("Niemiecki"), "de")
        self.ai_auto_language_combo.addItem(tr("Polski"), "pl")
        self.ai_auto_language_combo.addItem(tr("Angielski"), "en")
        self.ai_auto_language_combo.setToolTip(
            tr("Model pisze CAŁY mail w jednym języku. \"Auto\" pozwala AI samemu "
               "dopasować język do firmy (domyślnie niemiecki) - jeśli lokalny model "
               "(np. Ollama) miesza języki w treści, wybierz konkretny język zamiast Auto.")
        )
        ai_auto_row.addWidget(self.ai_auto_language_combo)
        ai_auto_row.addWidget(QLabel(tr('Próg AI:')))
        self.ai_auto_threshold_spin = QSpinBox()
        self.ai_auto_threshold_spin.setRange(0, 100)
        self.ai_auto_threshold_spin.setValue(50)
        self.ai_auto_threshold_spin.setSuffix(tr('%'))
        ai_auto_row.addWidget(self.ai_auto_threshold_spin)
        ai_auto_row.addWidget(QLabel(tr('Max leadów:')))
        self.ai_auto_max_spin = QSpinBox()
        self.ai_auto_max_spin.setRange(1, 1000)
        self.ai_auto_max_spin.setValue(100)
        ai_auto_row.addWidget(self.ai_auto_max_spin)
        ai_auto_layout.addLayout(ai_auto_row)

        ai_auto_btn_row = QHBoxLayout()
        self.ai_auto_start_btn = QPushButton(tr('🤖 START (w pełni automatycznie)'))
        self.ai_auto_start_btn.setStyleSheet("background: #5e2b8b; font-weight: bold; padding: 8px;")
        self.ai_auto_start_btn.clicked.connect(self.start_ai_auto_send)
        ai_auto_btn_row.addWidget(self.ai_auto_start_btn)
        self.ai_auto_stop_btn = QPushButton(tr('⏹️ STOP'))
        self.ai_auto_stop_btn.setStyleSheet("background: #8b0000; font-weight: bold;")
        self.ai_auto_stop_btn.clicked.connect(self.stop_ai_auto_send)
        self.ai_auto_stop_btn.setEnabled(False)
        ai_auto_btn_row.addWidget(self.ai_auto_stop_btn)
        ai_auto_layout.addLayout(ai_auto_btn_row)

        self.ai_auto_status = QLabel(tr('Gotowy'))
        ai_auto_layout.addWidget(self.ai_auto_status)
        self.ai_auto_counters = QLabel(tr('Przetworzono: 0 | Wysłano: 0 | Pominięto: 0 | Błędów: 0'))
        ai_auto_layout.addWidget(self.ai_auto_counters)
        layout.addWidget(ai_auto_group)

        self.ai_auto_worker = None

        self.send_progress = QProgressBar()
        layout.addWidget(self.send_progress)

        self.send_status = QLabel(tr('Gotowy do wysyłki'))
        layout.addWidget(self.send_status)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(container)
        self.send_tab.layout().addWidget(scroll)

    def refresh_send_list(self):
        self.send_list.clear()
        rows = db.get_leads()
        wyslane = db.get_wyslano_emails()
        for row in rows:
            lead_id, firma, _kontakt, email, *_rest, status, _wyslano = row
            text = f"{firma} | {email}"
            already_sent = status == 'wysłano' or email in wyslane
            if already_sent:
                text += " ✅"
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, {"id": lead_id, "firma": firma, "email": email, "status": status})
            if already_sent:
                item.setBackground(QColor(*COLOR_SENT_LIST))
            self.send_list.addItem(item)

    def select_new_only(self):
        wyslane = db.get_wyslano_emails()
        for i in range(self.send_list.count()):
            item = self.send_list.item(i)
            data = item.data(Qt.UserRole)
            if data and data.get('email') not in wyslane and data.get('status') != 'wysłano':
                item.setSelected(True)

    def preview_template(self):
        """Wyświetl podgląd wiadomości w oknie dialogowym."""
        try:
            dane = {'firma': 'Przykładowa Firma GmbH', 'kontakt': 'Jan Kowalski',
                    'email': 'kontakt@przykladowa-firma.de', 'adres': 'Musterstraße 1, 10115 Berlin',
                    'telefon': '+49 30 1234567', 'id': '123'}
            dane.update(get_company_info())
            for i in range(self.send_list.count()):
                item = self.send_list.item(i)
                if item.isSelected():
                    data = item.data(Qt.UserRole)
                    row = db.get_lead_by_id(data['id'])
                    if row:
                        dane.update({'firma': row[1] or dane['firma'], 'kontakt': row[2] or dane['kontakt'],
                                'email': row[3] or dane['email'], 'adres': row[4] or dane['adres'],
                                'telefon': row[5] or dane['telefon'], 'id': str(row[0])})
                    break

            szablon = self.szablon_edit.toPlainText().strip()
            temat = self.temat_edit.text().strip()
            if not szablon or not temat:
                QMessageBox.warning(self, tr('Błąd'), tr('Wpisz szablon i temat!'))
                return

            # Przetwórz zmienne, a potem spintax
            temat_final = SendWorker.resolve_spintax(SendWorker.parse_zmienne(temat, dane))
            tresc_final = SendWorker.resolve_spintax(SendWorker.parse_zmienne(szablon, dane))

            # Otwórz dialog
            dlg = QDialog(self)
            dlg.setWindowTitle(tr('👁 Podgląd wiadomości'))
            dlg.setGeometry(100, 100, 900, 600)
            lay = QVBoxLayout(dlg)
            lay.addWidget(QLabel(f"Do: {dane['email']}"))
            lay.addWidget(QLabel(tr('Temat:')))
            te = QTextEdit()
            te.setPlainText(temat_final)
            te.setReadOnly(True)
            te.setMaximumHeight(80)
            lay.addWidget(te)
            lay.addWidget(QLabel(tr('Treść:')))
            be = QTextEdit()
            be.setPlainText(tresc_final)
            be.setReadOnly(True)
            lay.addWidget(be, 1)
            btn = QPushButton(tr('Zamknij'))
            btn.clicked.connect(dlg.accept)
            lay.addWidget(btn)
            dlg.exec()
        except Exception as e:
            QMessageBox.critical(self, tr('Błąd'), f"Błąd w podglądzie: {str(e)}")
            logger.exception("Błąd w preview_template")

    def start_send(self, limit):
        min_score = self.lead_score_min_spin.value()
        allow_unscored = self.lead_score_allow_unscored.isChecked() if hasattr(self, 'lead_score_allow_unscored') else True
        wyslane = db.get_wyslano_emails()
        selected = []
        skipped_by_score = 0
        for i in range(self.send_list.count()):
            item = self.send_list.item(i)
            if item.isSelected():
                data = item.data(Qt.UserRole)
                if data and data.get('email') not in wyslane and data.get('status') != 'wysłano':
                    # Sprawdź Lead Score
                    lead_row = db.get_lead_by_id(data['id'])
                    if lead_row:
                        lead_score = lead_row[6] if len(lead_row) > 6 else -1
                        if lead_score == -1:
                            if allow_unscored:
                                selected.append(data)
                            else:
                                skipped_by_score += 1
                        elif lead_score >= min_score:
                            selected.append(data)
                        else:
                            skipped_by_score += 1
                    else:
                        selected.append(data)

        if not selected:
            msg = "Zaznacz leady do wysyłki (z odpowiednim score'em)!"
            if skipped_by_score:
                msg += f"\n({skipped_by_score} pominiętych z powodu progu punktacji AI)"
            QMessageBox.warning(self, tr('Błąd'), msg)
            return

        if limit < 9999:
            selected = selected[:limit]

        settings = get_current_profile_settings()
        gmail_user = settings.get("gmail_user", "")
        gmail_pass = settings.get("gmail_password", "")
        if not gmail_user or not gmail_pass:
            QMessageBox.warning(self, tr('Błąd'), tr('Skonfiguruj Gmail w ustawieniach profilu!'))
            return

        smtp_host = settings.get("smtp_host", SMTP_RELAY_HOST)
        smtp_port = settings.get("smtp_port", SMTP_RELAY_PORT)

        ok, msg = test_gmail_connection(gmail_user, gmail_pass, smtp_host, smtp_port)
        if not ok:
            QMessageBox.warning(self, tr('Błąd SMTP'), f"Nie można połączyć: {msg}")
            return

        leads = []
        for s in selected:
            row = db.get_lead_by_id(s['id'])
            if row:
                leads.append({
                    'id': row[0], 'firma': row[1] or '', 'kontakt': row[2] or '',
                    'email': row[3] or '', 'adres': row[4] or '', 'telefon': row[5] or ''
                })

        if not leads:
            QMessageBox.warning(self, tr('Błąd'), tr('Brak leadów!'))
            return

        dzienny = settings.get("dzienny_limit", SESSION_HARD_CAP)
        custom_delay = settings.get("custom_send_delay", CUSTOM_SEND_DELAY_DEFAULT)
        custom_cap = settings.get("custom_session_cap", CUSTOM_SESSION_CAP_DEFAULT)

        if len(leads) > dzienny:
            leads = leads[:dzienny]

        send_delay = get_send_delay(smtp_host, custom_delay)
        est_minutes = round(len(leads) * send_delay / 60)
        reply = QMessageBox.question(
            self, tr('Potwierdzenie'),
            f"Wysłać {len(leads)} wiadomości przez {smtp_host}?\n"
            f"Szacowany czas: ok. {est_minutes} min.",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.No:
            return

        szablon = self.szablon_edit.toPlainText().strip()
        temat = self.temat_edit.text().strip()
        if not szablon or not temat:
            QMessageBox.warning(self, tr('Błąd'), tr('Wpisz szablon i temat!'))
            return

        html = self.html_check.isChecked()
        attachments = [p.strip() for p in self.attachments_edit.text().split(',') if p.strip()]
        verify_mx = self.mx_check.isChecked()
        smime_sign = self.smime_check.isChecked()

        use_rotation = settings.get("account_rotation_enabled", False)
        rotator = None
        if use_rotation:
            accs = db.get_smtp_accounts()
            if accs:
                rotator = SMTPAccountRotator(accs)
                rotator.set_max_per_account(settings.get("rotation_max_per_account", 1000))

        self.send_worker = SendWorker(
            leads, szablon, temat, gmail_user, gmail_pass,
            smtp_host, smtp_port, dzienny, custom_delay, custom_cap,
            html=html, attachments=attachments,
            verify_mx=verify_mx, check_blacklist=True,
            smime_sign=smime_sign,
            personalized_attachments={},
            use_account_rotation=use_rotation,
            rotator=rotator,
            dry_run=self.dry_run_check.isChecked()
        )
        self.send_worker.progress.connect(self.update_send_progress)
        self.send_worker.status.connect(self.send_status.setText)
        self.send_worker.lead_done.connect(self.on_send_done)
        self.send_worker.finished.connect(self.on_send_finished)
        self.send_worker.error.connect(lambda e: self.send_status.setText(f"❌ {e}"))

        self.send_progress.setRange(0, len(leads))
        self.send_progress.setValue(0)
        self.send_worker.start()
        self.stop_send_btn.setEnabled(True)
        self.send_status.setText(tr('⏳ Wysyłanie...'))

    def stop_send(self):
        if self.send_worker:
            self.send_worker.stop()
            self.send_status.setText(tr('⏹️ Zatrzymywanie...'))

    def update_send_progress(self, current, total):
        self.send_progress.setValue(current)

    def on_send_done(self, lead):
        self.refresh_send_list()
        self.refresh_leads()

    def on_send_finished(self):
        self.stop_send_btn.setEnabled(False)
        self.refresh_send_list()
        self.refresh_leads()
        QMessageBox.information(self, tr('OK'), tr('Wysyłka zakończona!'))

    def start_ai_auto_send(self):
        selected = []
        for i in range(self.send_list.count()):
            item = self.send_list.item(i)
            if item.isSelected():
                data = item.data(Qt.UserRole)
                if data:
                    selected.append(data)

        if not selected:
            QMessageBox.warning(self, tr('Błąd'), tr('Zaznacz leady na liście powyżej (te same co do zwykłej wysyłki)!'))
            return

        max_count = self.ai_auto_max_spin.value()
        selected = selected[:max_count]

        settings = get_current_profile_settings()
        gmail_user = settings.get("gmail_user", "")
        gmail_pass = settings.get("gmail_password", "")
        if not gmail_user or not gmail_pass:
            QMessageBox.warning(self, tr('Błąd'), tr('Skonfiguruj Gmail/SMTP w ustawieniach profilu!'))
            return
        if not settings.get("company_offer_description", "").strip():
            QMessageBox.warning(
                self, tr('Błąd'),
                tr('Uzupełnij pole "Czym się zajmujemy" w ⚙️ Ustawienia -> Dane firmy - AI potrzebuje wiedzieć co oferujecie, żeby napisać sensowny mail.')
            )
            return

        smtp_host = settings.get("smtp_host", SMTP_RELAY_HOST)
        smtp_port = settings.get("smtp_port", SMTP_RELAY_PORT)
        ok, msg = test_gmail_connection(gmail_user, gmail_pass, smtp_host, smtp_port)
        if not ok:
            QMessageBox.warning(self, tr('Błąd SMTP'), f"Nie można połączyć: {msg}")
            return

        leads = []
        for s in selected:
            row = db.get_lead_by_id(s['id'])
            if row:
                leads.append({
                    'id': row[0], 'firma': row[1] or '', 'kontakt': row[2] or '',
                    'email': row[3] or '', 'adres': row[4] or '', 'telefon': row[5] or '',
                    'website': row[7] if len(row) > 7 else '',
                })
        if not leads:
            QMessageBox.warning(self, tr('Błąd'), tr('Brak leadów!'))
            return

        reply = QMessageBox.question(
            self, tr('Potwierdzenie'),
            f"AI przeczyta stronę i oceni {len(leads)} leadów, po czym BEZ dalszego "
            f"potwierdzania napisze i wyśle maile do tych, które przejdą próg "
            f"({self.ai_auto_threshold_spin.value()}%). To może chwilę potrwać "
            f"(czytanie stron + generowanie treści dla każdego leada). Kontynuować?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.No:
            return

        dzienny = settings.get("dzienny_limit", SESSION_HARD_CAP)
        custom_delay = settings.get("custom_send_delay", CUSTOM_SEND_DELAY_DEFAULT)
        custom_cap = settings.get("custom_session_cap", CUSTOM_SESSION_CAP_DEFAULT)
        html = self.html_check.isChecked()
        attachments = [p.strip() for p in self.attachments_edit.text().split(',') if p.strip()]
        verify_mx = self.mx_check.isChecked()
        smime_sign = self.smime_check.isChecked()

        use_rotation = settings.get("account_rotation_enabled", False)
        rotator = None
        if use_rotation:
            accs = db.get_smtp_accounts()
            if accs:
                rotator = SMTPAccountRotator(accs)
                rotator.set_max_per_account(settings.get("rotation_max_per_account", 1000))

        self.ai_auto_worker = AIAutoSendWorker(
            leads, gmail_user, gmail_pass, smtp_host, smtp_port,
            session_cap=dzienny, custom_delay=custom_delay, custom_cap=custom_cap,
            industry=self.ai_auto_industry_edit.text().strip(),
            ai_scoring_threshold=self.ai_auto_threshold_spin.value(),
            max_count=max_count,
            email_language=self.ai_auto_language_combo.currentData(),
            html=html, attachments=attachments,
            verify_mx=verify_mx, check_blacklist=True,
            smime_sign=smime_sign,
            use_account_rotation=use_rotation, rotator=rotator,
            dry_run=self.dry_run_check.isChecked(),
        )
        self.ai_auto_worker.status.connect(self.ai_auto_status.setText)
        self.ai_auto_worker.counters.connect(self.on_ai_auto_counters)
        self.ai_auto_worker.lead_processed.connect(lambda _l: (self.refresh_leads(), self.refresh_send_list()))
        self.ai_auto_worker.error.connect(lambda e: self.ai_auto_status.setText(f"❌ {e}"))
        self.ai_auto_worker.finished.connect(self.on_ai_auto_finished)
        self.ai_auto_worker.start()

        self.ai_auto_start_btn.setEnabled(False)
        self.ai_auto_stop_btn.setEnabled(True)
        self.ai_auto_status.setText(tr('⏳ Startuję (czytam pierwszą stronę)...'))

    def stop_ai_auto_send(self):
        if self.ai_auto_worker:
            self.ai_auto_worker.stop()
            self.ai_auto_status.setText(tr('⏹️ Zatrzymywanie...'))

    def on_ai_auto_counters(self, processed, sent, skipped, errors):
        self.ai_auto_counters.setText(
            f"Przetworzono: {processed} | Wysłano: {sent} | Pominięto: {skipped} | Błędów: {errors}"
        )

    def on_ai_auto_finished(self):
        self.ai_auto_start_btn.setEnabled(True)
        self.ai_auto_stop_btn.setEnabled(False)
        self.refresh_leads()
        self.refresh_send_list()
        QMessageBox.information(self, tr('OK'), tr('AI Auto-Send zakończony!'))

    # ------------------------------------------------------------------
    # ZAKŁADKA: USTAWIENIA
    # ------------------------------------------------------------------
    def build_settings_tab(self):
        subtabs = QTabWidget()

        basic_page = QWidget()
        layout = QVBoxLayout(basic_page)

        intro = QLabel(
            tr('To są ustawienia potrzebne do wysyłki: Twoja poczta, dane firmy do stopki i ile wiadomości wysyłać. Rzadziej używane opcje są w zakładce „Zaawansowane”.')
        )
        intro.setWordWrap(True)
        intro.setStyleSheet("color: #a6adc8; font-size: 12px;")
        layout.addWidget(intro)

        # Język aplikacji
        lang_group = QGroupBox(tr("🌐 Język aplikacji"))
        lang_layout = QHBoxLayout(lang_group)
        lang_layout.addWidget(QLabel(tr("Język interfejsu:")))
        self.language_combo = QComboBox()
        for code in SUPPORTED_LANGUAGES:
            self.language_combo.addItem(f"{LANGUAGE_FLAGS[code]} {LANGUAGE_NAMES[code]}", code)
        current_idx = SUPPORTED_LANGUAGES.index(get_language()) if get_language() in SUPPORTED_LANGUAGES else 0
        self.language_combo.setCurrentIndex(current_idx)
        self.language_combo.currentIndexChanged.connect(self._on_language_changed)
        lang_layout.addWidget(self.language_combo)
        lang_hint = QLabel(tr("ℹ️ Zmiana języka wymaga ponownego uruchomienia aplikacji."))
        lang_hint.setStyleSheet("color: #a6adc8; font-size: 11px;")
        lang_layout.addWidget(lang_hint)
        lang_layout.addStretch()
        layout.addWidget(lang_group)

        # Konto Gmail
        group = QGroupBox(tr('🔐 Konto Gmail'))
        form = QFormLayout(group)
        self.gmail_user = QLineEdit()
        form.addRow(tr('Adres Gmail:'), self.gmail_user)
        self.gmail_pass = QLineEdit()
        self.gmail_pass.setEchoMode(QLineEdit.Password)
        form.addRow(tr('Hasło aplikacji:'), self.gmail_pass)
        gmail_hint = QLabel(
            tr('ℹ️ To musi być "hasło aplikacji" Google (nie zwykłe hasło do konta) - wygenerujesz je w ustawieniach bezpieczeństwa konta Google.')
        )
        gmail_hint.setWordWrap(True)
        gmail_hint.setStyleSheet("color: #a6adc8; font-size: 11px;")
        form.addRow(gmail_hint)
        layout.addWidget(group)

        # Dane firmy (używane w szablonie - {company_name} itd.)
        company_group = QGroupBox(tr('🏢 Dane firmy (do stopki wiadomości)'))
        company_form = QFormLayout(company_group)
        self.company_name_edit = QLineEdit()
        self.company_name_edit.setPlaceholderText(tr('np. Twoja Firma Sp. z o.o.'))
        company_form.addRow(tr('Nazwa firmy:'), self.company_name_edit)
        self.company_address_edit = QLineEdit()
        self.company_address_edit.setPlaceholderText(tr('np. Ul. Przykładowa 1, 00-001 Miasto'))
        company_form.addRow(tr('Adres:'), self.company_address_edit)
        self.company_phone_edit = QLineEdit()
        self.company_phone_edit.setPlaceholderText(tr('np. +48 123 456 789'))
        company_form.addRow(tr('Telefon:'), self.company_phone_edit)
        self.company_email_edit = QLineEdit()
        self.company_email_edit.setPlaceholderText(tr('np. kontakt@twojafirma.pl'))
        company_form.addRow(tr('E-mail firmy:'), self.company_email_edit)
        self.company_website_edit = QLineEdit()
        self.company_website_edit.setPlaceholderText(tr('np. https://twojafirma.pl'))
        company_form.addRow(tr('Strona WWW:'), self.company_website_edit)
        self.company_offer_edit = QLineEdit()
        self.company_offer_edit.setPlaceholderText(
            tr('np. Tworzymy strony internetowe i sklepy online dla małych i średnich firm')
        )
        company_form.addRow(tr('Czym się zajmujemy (do AI):'), self.company_offer_edit)
        company_hint = QLabel(
            tr('ℹ️ Te dane pojawiają się automatycznie w stopce wiadomości - nie musisz ich wpisywać ręcznie w szablonie. Pole "Czym się zajmujemy" AI wykorzystuje do przedstawienia Waszej firmy w automatycznie pisanych mailach (sekcja 🤖 AI Auto-Send w zakładce Wysyłka) - opisz to prosto i konkretnie, AI nie będzie tego zmyślać.')
        )
        company_hint.setWordWrap(True)
        company_hint.setStyleSheet("color: #a6adc8; font-size: 11px;")
        company_form.addRow(company_hint)
        layout.addWidget(company_group)

        # Limity
        group2 = QGroupBox(tr('📊 Limity wysyłki'))
        form2 = QFormLayout(group2)
        self.dzienny_limit = QComboBox()
        form2.addRow(tr('Limit sesji:'), self.dzienny_limit)
        self.limit_pace_hint = QLabel()
        self.limit_pace_hint.setWordWrap(True)
        self.limit_pace_hint.setStyleSheet("color: #a6adc8; font-size: 11px;")
        form2.addRow(self.limit_pace_hint)
        layout.addWidget(group2)

        # Blacklist
        bl_group = QGroupBox(tr('🚫 Blacklist (adresy wykluczone z wysyłki)'))
        bl_layout = QVBoxLayout(bl_group)
        self.blacklist_list = QListWidget()
        bl_layout.addWidget(self.blacklist_list)
        bl_btn_row = QHBoxLayout()
        btn_refresh_bl = QPushButton(tr('🔄 Odśwież'))
        btn_refresh_bl.clicked.connect(self.refresh_blacklist)
        bl_btn_row.addWidget(btn_refresh_bl)
        btn_remove_bl = QPushButton(tr('🗑 Usuń zaznaczone'))
        btn_remove_bl.clicked.connect(self.remove_from_blacklist)
        bl_btn_row.addWidget(btn_remove_bl)
        btn_import_bl = QPushButton(tr('📥 Importuj z pliku'))
        btn_import_bl.clicked.connect(self.import_blacklist)
        bl_btn_row.addWidget(btn_import_bl)
        bl_layout.addLayout(bl_btn_row)
        layout.addWidget(bl_group)
        layout.addStretch()

        basic_scroll = QScrollArea()
        basic_scroll.setWidgetResizable(True)
        basic_scroll.setWidget(basic_page)
        subtabs.addTab(basic_scroll, tr('🟢 Podstawowe'))

        # ================================================================
        # PODZAKŁADKA: AI
        # ================================================================
        ai_page = QWidget()
        ai_page_layout = QVBoxLayout(ai_page)

        ai_intro = QLabel(
            tr('Tu ustawiasz tylko próg oceny leadów przez AI. Wybór dostawcy AI (OpenAI/Gemini/Ollama itd.) i klucze API ustawiasz w zakładce 📁 Kampania → 🤖 Asystent AI.')
        )
        ai_intro.setWordWrap(True)
        ai_intro.setStyleSheet("color: #a6adc8; font-size: 12px;")
        ai_page_layout.addWidget(ai_intro)

        group5 = QGroupBox(tr('🤖 AI Lead Scoring'))
        form5 = QFormLayout(group5)
        self.ai_scoring_enabled = QCheckBox(tr('Włącz AI Lead Scoring'))
        form5.addRow(self.ai_scoring_enabled)
        self.ai_scoring_threshold = QSpinBox()
        self.ai_scoring_threshold.setRange(0, 100)
        self.ai_scoring_threshold.setValue(50)
        self.ai_scoring_threshold.setSuffix(tr('%'))
        self.ai_scoring_threshold.valueChanged.connect(self._on_ai_scoring_threshold_changed)
        form5.addRow(tr('Próg punktacji minimalnej:'), self.ai_scoring_threshold)
        ai_scoring_hint = QLabel(
            tr('ℹ️ Gdy włączone, AI ocenia każdego leada i pomija w wysyłce te poniżej progu.')
        )
        ai_scoring_hint.setWordWrap(True)
        ai_scoring_hint.setStyleSheet("color: #a6adc8; font-size: 11px;")
        form5.addRow(ai_scoring_hint)
        ai_page_layout.addWidget(group5)
        ai_page_layout.addStretch()

        ai_scroll = QScrollArea()
        ai_scroll.setWidgetResizable(True)
        ai_scroll.setWidget(ai_page)
        subtabs.addTab(ai_scroll, tr('🤖 AI'))

        # ================================================================
        # PODZAKŁADKA: ZAAWANSOWANE
        # ================================================================
        adv_page = QWidget()
        adv_layout = QVBoxLayout(adv_page)

        adv_intro = QLabel(
            tr('⚠️ Te opcje są dla zaawansowanych przypadków (własny serwer poczty, wiele kont, proxy, podpis cyfrowy). Jeśli nie wiesz do czego to służy, prawdopodobnie nie musisz tego ruszać - domyślne ustawienia działają.')
        )
        adv_intro.setWordWrap(True)
        adv_intro.setStyleSheet("color: #f9a825; font-size: 12px;")
        adv_layout.addWidget(adv_intro)

        # Serwer SMTP
        group3 = QGroupBox(tr('🌐 Serwer SMTP'))
        form3 = QFormLayout(group3)
        self.smtp_mode = QComboBox()
        self.smtp_mode.addItems([
            "Google Workspace SMTP Relay",
            "Zwykły Gmail SMTP",
            "Inny dostawca",
        ])
        self.smtp_mode.currentIndexChanged.connect(self._apply_smtp_mode)
        form3.addRow(tr('Tryb:'), self.smtp_mode)
        self.smtp_host_edit = QLineEdit()
        form3.addRow(tr('Host:'), self.smtp_host_edit)
        self.smtp_port_edit = QLineEdit()
        form3.addRow(tr('Port:'), self.smtp_port_edit)

        self.custom_delay_spin = QDoubleSpinBox()
        self.custom_delay_spin.setRange(CUSTOM_SEND_DELAY_MIN, 30.0)
        self.custom_delay_spin.setSingleStep(0.5)
        self.custom_delay_spin.setSuffix(tr(' s/wiadomość'))
        self.custom_delay_spin.setValue(CUSTOM_SEND_DELAY_DEFAULT)
        self.custom_delay_spin.valueChanged.connect(self._on_custom_smtp_changed)
        form3.addRow(tr('Własne tempo:'), self.custom_delay_spin)

        self.custom_cap_spin = QSpinBox()
        self.custom_cap_spin.setRange(10, CUSTOM_SESSION_CAP_MAX)
        self.custom_cap_spin.setSingleStep(50)
        self.custom_cap_spin.setValue(CUSTOM_SESSION_CAP_DEFAULT)
        self.custom_cap_spin.valueChanged.connect(self._on_custom_smtp_changed)
        form3.addRow(tr('Własny limit sesji:'), self.custom_cap_spin)
        adv_layout.addWidget(group3)

        # Rotacja kont - z autodetekcją
        rot_group = QGroupBox(tr('🔄 Rotacja kont SMTP (wysyłka z wielu skrzynek)'))
        rot_layout = QVBoxLayout(rot_group)

        self.rotation_enabled = QCheckBox(tr('Włącz rotację kont'))
        rot_layout.addWidget(self.rotation_enabled)

        self.rotation_max_spin = QSpinBox()
        self.rotation_max_spin.setRange(100, 5000)
        self.rotation_max_spin.setValue(1000)
        hbox_max = QHBoxLayout()
        hbox_max.addWidget(QLabel(tr('Maks. wiadomości na konto:')))
        hbox_max.addWidget(self.rotation_max_spin)
        hbox_max.addStretch()
        rot_layout.addLayout(hbox_max)

        rot_layout.addWidget(QLabel(tr('Dodane konta:')))
        self.accounts_list = QListWidget()
        rot_layout.addWidget(self.accounts_list)

        # FORMULARZ DODAWANIA KONTA
        add_group = QGroupBox(tr('➕ Dodaj nowe konto'))
        add_form = QVBoxLayout(add_group)

        row_email = QHBoxLayout()
        self.account_email = QLineEdit()
        self.account_email.setPlaceholderText(tr('adres@example.com'))
        self.account_email.editingFinished.connect(self._auto_fill_smtp)
        row_email.addWidget(QLabel(tr('E-mail:')))
        row_email.addWidget(self.account_email)

        self.account_pass = QLineEdit()
        self.account_pass.setPlaceholderText(tr('hasło / hasło aplikacji'))
        self.account_pass.setEchoMode(QLineEdit.Password)
        row_email.addWidget(QLabel(tr('Hasło:')))
        row_email.addWidget(self.account_pass)
        add_form.addLayout(row_email)

        row_host = QHBoxLayout()
        self.account_host = QLineEdit()
        self.account_host.setPlaceholderText(tr('smtp.domena.com'))
        row_host.addWidget(QLabel(tr('Host:')))
        row_host.addWidget(self.account_host)

        self.account_port = QLineEdit()
        self.account_port.setPlaceholderText(tr('587'))
        self.account_port.setText(tr('587'))
        row_host.addWidget(QLabel(tr('Port:')))
        row_host.addWidget(self.account_port)
        add_form.addLayout(row_host)

        btn_add_account = QPushButton(tr('➕ Dodaj konto'))
        btn_add_account.clicked.connect(self.add_smtp_account)
        add_form.addWidget(btn_add_account)

        rot_layout.addWidget(add_group)

        btn_del_account = QPushButton(tr('🗑 Usuń zaznaczone konto'))
        btn_del_account.clicked.connect(self.remove_smtp_account)
        rot_layout.addWidget(btn_del_account)

        adv_layout.addWidget(rot_group)

        # IMAP
        imap_group = QGroupBox(tr('📨 Monitorowanie zwrotów (IMAP)'))
        imap_layout = QVBoxLayout(imap_group)
        self.imap_enabled = QCheckBox(tr('Włącz monitorowanie zwrotów'))
        imap_layout.addWidget(self.imap_enabled)
        self.imap_server_edit = QLineEdit("imap.gmail.com")
        imap_layout.addWidget(QLabel(tr('Serwer IMAP:')))
        imap_layout.addWidget(self.imap_server_edit)
        self.imap_user_edit = QLineEdit()
        imap_layout.addWidget(QLabel(tr('Użytkownik IMAP:')))
        imap_layout.addWidget(self.imap_user_edit)
        self.imap_pass_edit = QLineEdit()
        self.imap_pass_edit.setEchoMode(QLineEdit.Password)
        imap_layout.addWidget(QLabel(tr('Hasło IMAP:')))
        imap_layout.addWidget(self.imap_pass_edit)
        adv_layout.addWidget(imap_group)

        # S/MIME
        smime_group = QGroupBox(tr('🔏 S/MIME (podpis cyfrowy wiadomości)'))
        smime_layout = QVBoxLayout(smime_group)
        self.smime_enabled = QCheckBox(tr('Włącz podpisywanie S/MIME'))
        smime_layout.addWidget(self.smime_enabled)
        btn_gen_cert = QPushButton(tr('🔄 Wygeneruj nowy certyfikat'))
        btn_gen_cert.clicked.connect(self.generate_smime_cert)
        smime_layout.addWidget(btn_gen_cert)
        adv_layout.addWidget(smime_group)

        # Proxy
        group4 = QGroupBox(tr('🌐 Proxy (wyszukiwanie)'))
        form4 = QFormLayout(group4)
        self.proxy_enabled_check = QCheckBox(tr('Włącz proxy'))
        form4.addRow(self.proxy_enabled_check)
        self.proxy_list_edit = QTextEdit()
        self.proxy_list_edit.setMaximumHeight(100)
        form4.addRow(tr('Lista proxy:'), self.proxy_list_edit)
        adv_layout.addWidget(group4)
        adv_layout.addStretch()

        adv_scroll = QScrollArea()
        adv_scroll.setWidgetResizable(True)
        adv_scroll.setWidget(adv_page)
        subtabs.addTab(adv_scroll, tr('⚙️ Zaawansowane'))

        # ================================================================
        # Przyciski akcji - zawsze widoczne, niezależnie od otwartej pod-zakładki
        # ================================================================
        self.settings_tab.layout().addWidget(subtabs)

        btn_row = QHBoxLayout()
        btn_save = QPushButton(tr('💾 Zapisz ustawienia profilu'))
        btn_save.clicked.connect(self.save_settings)
        btn_row.addWidget(btn_save)

        btn_test = QPushButton(tr('🔍 Test połączenia SMTP'))
        btn_test.clicked.connect(self.test_gmail_ui)
        btn_row.addWidget(btn_test)

        self.settings_tab.layout().addLayout(btn_row)

    # ------------------------------------------------------------------
    # METODY USTAWIEŃ
    # ------------------------------------------------------------------
    def load_settings(self):
        settings = get_current_profile_settings()

        self.gmail_user.setText(settings.get("gmail_user", ""))
        self.gmail_pass.setText(settings.get("gmail_password", ""))

        self.company_name_edit.setText(settings.get("company_name", ""))
        self.company_address_edit.setText(settings.get("company_address", ""))
        self.company_phone_edit.setText(settings.get("company_phone", ""))
        self.company_email_edit.setText(settings.get("company_email", ""))
        self.company_website_edit.setText(settings.get("company_website", ""))
        self.company_offer_edit.setText(settings.get("company_offer_description", ""))

        saved_host = settings.get("smtp_host", SMTP_RELAY_HOST)
        saved_port = settings.get("smtp_port", SMTP_RELAY_PORT)

        if saved_host == SMTP_RELAY_HOST:
            mode_index = 0
        elif saved_host == SMTP_FALLBACK_HOST:
            mode_index = 1
        else:
            mode_index = 2

        self.custom_delay_spin.setValue(settings.get("custom_send_delay", CUSTOM_SEND_DELAY_DEFAULT))
        self.custom_cap_spin.setValue(settings.get("custom_session_cap", CUSTOM_SESSION_CAP_DEFAULT))
        self.custom_delay_spin.setEnabled(mode_index == 2)
        self.custom_cap_spin.setEnabled(mode_index == 2)

        self.smtp_mode.blockSignals(True)
        self.smtp_mode.setCurrentIndex(mode_index)
        self.smtp_mode.blockSignals(False)
        self.smtp_host_edit.setText(saved_host)
        self.smtp_port_edit.setText(str(saved_port))

        default_limit = {0: SESSION_HARD_CAP, 1: GMAIL_FREE_SESSION_CAP_DEFAULT,
                          2: self.custom_cap_spin.value()}[mode_index]
        saved_limit = settings.get("dzienny_limit", default_limit)
        self._populate_session_cap_options(mode_index, preselect=saved_limit)

        self.proxy_enabled_check.setChecked(settings.get("proxy_enabled", False))
        self.proxy_list_edit.setPlainText(settings.get("proxy_list", ""))

        self.html_check.setChecked(settings.get("html_enabled", False))
        self.mx_check.setChecked(settings.get("mx_verify_enabled", False))
        self.smime_check.setChecked(settings.get("smime_enabled", False))
        self.attachments_edit.setText(settings.get("attachments", ""))

        self.rotation_enabled.setChecked(settings.get("account_rotation_enabled", False))
        self.rotation_max_spin.setValue(settings.get("rotation_max_per_account", 1000))

        # Wczytaj konta do listy
        self.accounts_data = []
        self.accounts_list.clear()
        accounts = db.get_smtp_accounts()
        for acc in accounts:
            if acc['enabled']:
                self.accounts_data.append(acc)
                self.accounts_list.addItem(f"{acc['user']}@{acc['host']}:{acc['port']}")

        self.imap_enabled.setChecked(settings.get("imap_enabled", False))
        self.imap_server_edit.setText(settings.get("imap_server", "imap.gmail.com"))
        self.imap_user_edit.setText(settings.get("imap_user", ""))
        self.imap_pass_edit.setText(settings.get("imap_password", ""))

        self.smime_enabled.setChecked(settings.get("smime_enabled", False))

        # Wczytaj ustawienia AI Scoring
        self.ai_scoring_enabled.setChecked(settings.get("ai_scoring_enabled", False))
        self.ai_scoring_threshold.setValue(settings.get("ai_scoring_threshold", 50))
        if hasattr(self, 'lead_score_min_spin'):
            self.lead_score_min_spin.blockSignals(True)
            self.lead_score_min_spin.setValue(settings.get("ai_scoring_threshold", 50))
            self.lead_score_min_spin.blockSignals(False)
        if hasattr(self, 'lead_score_allow_unscored'):
            self.lead_score_allow_unscored.blockSignals(True)
            self.lead_score_allow_unscored.setChecked(settings.get("lead_score_allow_unscored", True))
            self.lead_score_allow_unscored.blockSignals(False)

        # Wczytaj szablony i zapytania
        self.queries_edit.setPlainText(settings.get("last_queries", DEFAULT_QUERIES))
        self.locations_edit.setPlainText(settings.get("last_locations", DEFAULT_LOCATIONS))
        self.szablon_edit.setPlainText(settings.get("last_template", DEFAULT_TEMPLATE))
        self.temat_edit.setText(settings.get("last_subject", DEFAULT_SUBJECT))

    def save_settings(self):
        settings = {
            "gmail_user": self.gmail_user.text().strip(),
            "gmail_password": self.gmail_pass.text(),
            "company_name": self.company_name_edit.text().strip(),
            "company_address": self.company_address_edit.text().strip(),
            "company_phone": self.company_phone_edit.text().strip(),
            "company_email": self.company_email_edit.text().strip(),
            "company_website": self.company_website_edit.text().strip(),
            "company_offer_description": self.company_offer_edit.text().strip(),
            "smtp_host": self.smtp_host_edit.text().strip() or SMTP_RELAY_HOST,
            "smtp_port": int(self.smtp_port_edit.text().strip()) if self.smtp_port_edit.text().strip().isdigit() else SMTP_RELAY_PORT,
            "custom_send_delay": self.custom_delay_spin.value(),
            "custom_session_cap": self.custom_cap_spin.value(),
            "dzienny_limit": self.dzienny_limit.currentData(),
            "proxy_enabled": self.proxy_enabled_check.isChecked(),
            "proxy_list": self.proxy_list_edit.toPlainText().strip(),
            "html_enabled": self.html_check.isChecked(),
            "mx_verify_enabled": self.mx_check.isChecked(),
            "smime_enabled": self.smime_check.isChecked(),
            "attachments": self.attachments_edit.text().strip(),
            "account_rotation_enabled": self.rotation_enabled.isChecked(),
            "rotation_max_per_account": self.rotation_max_spin.value(),
            "imap_enabled": self.imap_enabled.isChecked(),
            "imap_server": self.imap_server_edit.text().strip(),
            "imap_user": self.imap_user_edit.text().strip(),
            "imap_password": self.imap_pass_edit.text(),
            "ai_scoring_enabled": self.ai_scoring_enabled.isChecked(),
            "ai_scoring_threshold": self.ai_scoring_threshold.value(),
            # Szablony i zapytania
            "last_queries": self.queries_edit.toPlainText(),
            "last_locations": self.locations_edit.toPlainText(),
            "last_template": self.szablon_edit.toPlainText(),
            "last_subject": self.temat_edit.text(),
        }
        update_current_profile_settings(settings)

        # Zapisz konta SMTP w bazie (osobno)
        db.save_smtp_accounts(self.accounts_data)

        # Uruchom/zatrzymaj monitor zwrotów
        if settings["imap_enabled"] and settings["imap_user"] and settings["imap_password"]:
            if not self.bounce_monitor or not self.bounce_monitor.is_alive():
                self.bounce_monitor = BounceMonitor(
                    settings["imap_user"],
                    settings["imap_password"],
                    settings["imap_server"],
                    on_bounce=lambda addr: (db.add_to_blacklist(addr, "bounce"), self.refresh_blacklist())
                )
                self.bounce_monitor.start()
        else:
            if self.bounce_monitor and self.bounce_monitor.is_alive():
                self.bounce_monitor.stop()
                self.bounce_monitor = None

        self._refresh_inbox_accounts()
        QMessageBox.information(self, tr('OK'), tr('Ustawienia profilu zapisane!'))

    def _apply_smtp_mode(self, index):
        if index == 0:
            self.smtp_host_edit.setText(SMTP_RELAY_HOST)
            self.smtp_port_edit.setText(str(SMTP_RELAY_PORT))
        elif index == 1:
            self.smtp_host_edit.setText(SMTP_FALLBACK_HOST)
            self.smtp_port_edit.setText(str(SMTP_FALLBACK_PORT))
        is_custom = (index == 2)
        self.custom_delay_spin.setEnabled(is_custom)
        self.custom_cap_spin.setEnabled(is_custom)
        self._populate_session_cap_options(index)

    def _on_custom_smtp_changed(self, _value):
        if self.smtp_mode.currentIndex() == 2:
            self._populate_session_cap_options(2)

    def _populate_session_cap_options(self, mode_index: int, preselect: Optional[int] = None):
        self.dzienny_limit.blockSignals(True)
        self.dzienny_limit.clear()
        if mode_index == 0:
            for value in SESSION_CAP_OPTIONS:
                label = f"{value:,} / sesję".replace(",", " ")
                if value == SESSION_HARD_CAP:
                    label += " (zalecane)"
                self.dzienny_limit.addItem(label, value)
            delay, unit = SEND_FIXED_DELAY, "Google Workspace Relay"
        elif mode_index == 1:
            for value in GMAIL_FREE_SESSION_CAP_OPTIONS:
                label = f"{value} / sesję"
                self.dzienny_limit.addItem(label, value)
            delay, unit = get_send_delay(SMTP_FALLBACK_HOST), "zwykły Gmail"
        else:
            cap = self.custom_cap_spin.value()
            self.dzienny_limit.addItem(f"{cap} / sesję", cap)
            delay, unit = self.custom_delay_spin.value(), "inny dostawca"

        idx = self.dzienny_limit.findData(preselect) if preselect is not None else -1
        self.dzienny_limit.setCurrentIndex(idx if idx >= 0 else 0)
        self.dzienny_limit.blockSignals(False)
        self.limit_pace_hint.setText(
            f"Tryb: {unit}. Tempo: {delay:g} s/wiadomość (~{round(60/delay)}/min)."
        )

    def test_gmail_ui(self):
        user = self.gmail_user.text().strip()
        password = self.gmail_pass.text().strip()
        host = self.smtp_host_edit.text().strip() or SMTP_RELAY_HOST
        try:
            port = int(self.smtp_port_edit.text().strip())
        except ValueError:
            port = SMTP_RELAY_PORT
        if not user or not password:
            QMessageBox.warning(self, tr('Błąd'), tr('Wpisz dane logowania!'))
            return

        ok, msg = test_gmail_connection(user, password, host, port)
        if ok:
            QMessageBox.information(self, tr('OK'), f"✅ Połączenie z {host}:{port} działa!")
        else:
            QMessageBox.warning(self, tr('Błąd'), f"❌ Nie można połączyć: {msg}")

    # ------------------------------------------------------------------
    # BLACKLIST
    # ------------------------------------------------------------------
    def refresh_blacklist(self):
        self.blacklist_list.clear()
        for email, reason, added in db.get_blacklist():
            self.blacklist_list.addItem(f"{email}  ({reason})  [{added[:10]}]")

    def remove_from_blacklist(self):
        for item in self.blacklist_list.selectedItems():
            email = item.text().split()[0]
            db.remove_from_blacklist(email)
        self.refresh_blacklist()

    def import_blacklist(self):
        path, _ = QFileDialog.getOpenFileName(self, "Wybierz plik z adresami", "", "TXT/CSV (*.txt *.csv)")
        if not path:
            return
        try:
            count = import_blacklist_from_file(path, "import")
            self.refresh_blacklist()
            QMessageBox.information(self, tr('OK'), f"Zaimportowano {count} adresów.")
        except Exception as e:
            QMessageBox.critical(self, tr('Błąd'), str(e))

    # ------------------------------------------------------------------
    # KONTA SMTP (rotacja)
    # ------------------------------------------------------------------
    def _auto_fill_smtp(self):
        email = self.account_email.text().strip()
        if not email or '@' not in email:
            return
        guessed = guess_smtp(email)
        if guessed:
            host, port = guessed
            self.account_host.setText(host)
            self.account_port.setText(str(port))
        else:
            domain = email.split('@')[1].lower()
            self.account_host.setText(f"smtp.{domain}")
            self.account_port.setText(tr('587'))

    def add_smtp_account(self):
        email = self.account_email.text().strip()
        password = self.account_pass.text().strip()
        host = self.account_host.text().strip()
        port_str = self.account_port.text().strip()

        if not email or not password or not host or not port_str:
            QMessageBox.warning(self, tr('Błąd'), tr('Wypełnij wszystkie pola.'))
            return

        try:
            port_int = int(port_str)
        except ValueError:
            QMessageBox.warning(self, tr('Błąd'), tr('Port musi być liczbą.'))
            return

        self.accounts_data.append({
            'user': email,
            'password': password,
            'host': host,
            'port': port_int,
            'enabled': True
        })
        self.accounts_list.addItem(f"{email}@{host}:{port_int}")
        self.account_email.clear()
        self.account_pass.clear()
        self.account_host.clear()
        self.account_port.setText(tr('587'))

    def remove_smtp_account(self):
        for item in self.accounts_list.selectedItems():
            row = self.accounts_list.row(item)
            self.accounts_list.takeItem(row)
            if row < len(self.accounts_data):
                del self.accounts_data[row]

    # ------------------------------------------------------------------
    # S/MIME
    # ------------------------------------------------------------------
    def generate_smime_cert(self):
        from core.smime import generate_smime_cert
        try:
            p12, finger = generate_smime_cert()
            QMessageBox.information(self, tr('OK'), f"Certyfikat wygenerowany.\nFingerprint: {finger}")
        except Exception as e:
            QMessageBox.critical(self, tr('Błąd'), str(e))

    # ------------------------------------------------------------------
    # ZAKŁADKA: HISTORIA
    # ------------------------------------------------------------------
    def build_history_tab(self):
        container = QWidget()
        layout = QVBoxLayout(container)

        self.history_table = QTableWidget()
        self.history_table.setColumnCount(5)
        self.history_table.setHorizontalHeaderLabels([tr('Data'), tr('Email'), tr('Status'), tr('Temat'), tr('Błąd')])
        layout.addWidget(self.history_table)

        btn_refresh = QPushButton(tr('🔄 Odśwież'))
        btn_refresh.clicked.connect(self.refresh_history)
        layout.addWidget(btn_refresh)

        btn_clear = QPushButton(tr('🗑 Wyczyść logi (starsze niż 30 dni)'))
        btn_clear.clicked.connect(self.clear_old_logs)
        layout.addWidget(btn_clear)

        if MATPLOTLIB_AVAILABLE:
            self.stats_frame = QWidget()
            stats_layout = QVBoxLayout(self.stats_frame)
            btn_stats = QPushButton(tr('📊 Odśwież wykres statystyk'))
            btn_stats.clicked.connect(self.refresh_stats)
            stats_layout.addWidget(btn_stats)
            self.stats_canvas = None
            stats_layout.addStretch()
            layout.addWidget(self.stats_frame)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(container)
        self.history_tab.layout().addWidget(scroll)

    def build_inbox_tab(self):
        container = QWidget()
        layout = QVBoxLayout(container)

        intro = QLabel(
            tr('Podgląd wiadomości przychodzących na konta e-mail przypisane do bieżącego profilu. Wybierz konto i folder, kliknij „Odśwież”.')
        )
        intro.setWordWrap(True)
        intro.setStyleSheet("color: #a6adc8; font-size: 12px;")
        layout.addWidget(intro)

        row = QHBoxLayout()
        row.addWidget(QLabel(tr('Konto:')))
        self.inbox_account_combo = QComboBox()
        self.inbox_account_combo.currentIndexChanged.connect(self._on_inbox_account_changed)
        row.addWidget(self.inbox_account_combo, stretch=2)

        row.addWidget(QLabel(tr('Folder:')))
        self.inbox_folder_combo = QComboBox()
        self.inbox_folder_combo.addItem("INBOX")
        self.inbox_folder_combo.setEditable(True)
        row.addWidget(self.inbox_folder_combo, stretch=1)

        row.addWidget(QLabel(tr('Serwer IMAP:')))
        self.inbox_server_edit = QLineEdit()
        row.addWidget(self.inbox_server_edit, stretch=1)

        btn_refresh_inbox = QPushButton(tr('🔄 Odśwież'))
        btn_refresh_inbox.clicked.connect(self.refresh_inbox)
        row.addWidget(btn_refresh_inbox)
        layout.addLayout(row)

        self.inbox_status = QLabel("")
        self.inbox_status.setWordWrap(True)
        self.inbox_status.setStyleSheet("color: #a6adc8; font-size: 11px;")
        layout.addWidget(self.inbox_status)

        self.inbox_table = QTableWidget()
        self.inbox_table.setColumnCount(3)
        self.inbox_table.setHorizontalHeaderLabels([tr('Od'), tr('Temat'), tr('Data')])
        self.inbox_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.inbox_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.inbox_table.itemSelectionChanged.connect(self._on_inbox_message_selected)
        layout.addWidget(self.inbox_table, stretch=1)

        action_row = QHBoxLayout()
        btn_reply = QPushButton(tr('✉️ Odpowiedz'))
        btn_reply.clicked.connect(self._reply_to_inbox_message)
        action_row.addWidget(btn_reply)

        self.btn_toggle_read = QPushButton(tr('👁 Oznacz jako (nie)przeczytane'))
        self.btn_toggle_read.clicked.connect(self._toggle_read_inbox_message)
        action_row.addWidget(self.btn_toggle_read)

        btn_delete_msg = QPushButton(tr('🗑 Usuń'))
        btn_delete_msg.clicked.connect(self._delete_inbox_message)
        action_row.addWidget(btn_delete_msg)
        layout.addLayout(action_row)

        layout.addWidget(QLabel(tr('Treść wiadomości:')))
        self.inbox_body = QTextEdit()
        self.inbox_body.setReadOnly(True)
        layout.addWidget(self.inbox_body, stretch=1)

        layout.addWidget(QLabel(tr('📎 Załączniki (podwójny klik, żeby zapisać):')))
        self.inbox_attachments_list = QListWidget()
        self.inbox_attachments_list.setMaximumHeight(80)
        self.inbox_attachments_list.itemDoubleClicked.connect(self._save_inbox_attachment)
        layout.addWidget(self.inbox_attachments_list)

        self.inbox_tab.layout().addWidget(container)
        self._current_full_message = None
        self._refresh_inbox_accounts()

    def _get_profile_email_accounts(self) -> List[Dict]:
        """Konta e-mail przypisane do bieżącego profilu: główne konto Gmail
        (z Ustawień) + wszystkie konta z rotacji SMTP."""
        accounts = []
        settings = get_current_profile_settings()
        primary_user = settings.get("gmail_user", "").strip()
        if primary_user:
            accounts.append({
                "label": f"{primary_user} (główne)",
                "email": primary_user,
                "password": settings.get("gmail_password", ""),
                "host": settings.get("smtp_host", ""),
            })
        for acc in db.get_smtp_accounts():
            if not acc.get("enabled", True) or not acc.get("user"):
                continue
            if any(a["email"] == acc["user"] for a in accounts):
                continue
            accounts.append({
                "label": f"{acc['user']} (rotacja)",
                "email": acc["user"],
                "password": acc.get("password", ""),
                "host": acc.get("host", ""),
            })
        return accounts

    def _refresh_inbox_accounts(self):
        self.inbox_account_combo.blockSignals(True)
        self.inbox_account_combo.clear()
        self._inbox_accounts = self._get_profile_email_accounts()
        for acc in self._inbox_accounts:
            self.inbox_account_combo.addItem(acc["label"])
        self.inbox_account_combo.blockSignals(False)
        if self._inbox_accounts:
            self._on_inbox_account_changed(0)
        else:
            self.inbox_status.setText(
                tr('ℹ️ Brak skonfigurowanego konta e-mail dla tego profilu - uzupełnij Gmail w zakładce ⚙️ Ustawienia.')
            )

    def _on_inbox_account_changed(self, index: int):
        if index < 0 or index >= len(getattr(self, "_inbox_accounts", [])):
            return
        acc = self._inbox_accounts[index]
        self.inbox_server_edit.setText(guess_imap_server(acc["email"], acc.get("host", "")))
        if not acc.get("password"):
            return
        self._folder_worker = FolderListWorker(
            acc["email"], acc["password"], self.inbox_server_edit.text().strip()
        )
        self._folder_worker.finished_ok.connect(self._on_folders_fetched)
        self._folder_worker.finished_error.connect(lambda err: None)  # cicho - INBOX zawsze zadziała
        self._folder_worker.start()

    def _on_folders_fetched(self, folders: list):
        self._inbox_folders = folders
        current = self.inbox_folder_combo.currentText() or "INBOX"
        self.inbox_folder_combo.blockSignals(True)
        self.inbox_folder_combo.clear()
        for f in folders:
            self.inbox_folder_combo.addItem(f)
        idx = self.inbox_folder_combo.findText(current)
        self.inbox_folder_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self.inbox_folder_combo.blockSignals(False)

    def _current_inbox_account(self) -> Optional[Dict]:
        idx = self.inbox_account_combo.currentIndex()
        if idx < 0 or idx >= len(getattr(self, "_inbox_accounts", [])):
            return None
        return self._inbox_accounts[idx]

    def refresh_inbox(self):
        acc = self._current_inbox_account()
        if not acc:
            QMessageBox.warning(self, tr('Błąd'), tr('Wybierz konto e-mail.'))
            return
        imap_server = self.inbox_server_edit.text().strip()
        folder = self.inbox_folder_combo.currentText().strip() or "INBOX"
        if not acc.get("password"):
            QMessageBox.warning(
                self, tr('Błąd'),
                tr('To konto nie ma zapisanego hasła aplikacji - uzupełnij je w Ustawieniach.')
            )
            return

        self.inbox_status.setText(tr('⏳ Łączenie ze skrzynką...'))
        self.inbox_table.setRowCount(0)
        self.inbox_body.clear()
        self.inbox_attachments_list.clear()

        self._inbox_worker = InboxFetchWorker(acc["email"], acc["password"], imap_server, folder)
        self._inbox_worker.finished_ok.connect(self._on_inbox_fetched)
        self._inbox_worker.finished_error.connect(
            lambda err: self.inbox_status.setText(f"❌ Nie można pobrać wiadomości: {err}")
        )
        self._inbox_worker.start()

    def _on_inbox_fetched(self, messages: list):
        self._inbox_messages = messages
        self.inbox_table.setRowCount(len(messages))
        for i, msg in enumerate(messages):
            self.inbox_table.setItem(i, 0, QTableWidgetItem(msg.sender))
            item_subject = QTableWidgetItem(msg.subject)
            if not msg.seen:
                item_subject.setForeground(QColor(*COLOR_OK))
            self.inbox_table.setItem(i, 1, item_subject)
            self.inbox_table.setItem(i, 2, QTableWidgetItem(msg.date))
        self.inbox_table.resizeColumnsToContents()
        self.inbox_status.setText(f"✅ Wczytano {len(messages)} wiadomości.")

    def _selected_inbox_message(self):
        rows = self.inbox_table.selectionModel().selectedRows()
        if not rows or not getattr(self, "_inbox_messages", None):
            return None
        row_index = rows[0].row()
        if row_index >= len(self._inbox_messages):
            return None
        return self._inbox_messages[row_index]

    def _on_inbox_message_selected(self):
        msg = self._selected_inbox_message()
        if not msg:
            return
        acc = self._current_inbox_account()
        if not acc:
            return
        imap_server = self.inbox_server_edit.text().strip()
        folder = self.inbox_folder_combo.currentText().strip() or "INBOX"

        self.inbox_body.setPlainText(tr('⏳ Wczytywanie treści...'))
        self.inbox_attachments_list.clear()
        self._full_worker = MessageFullWorker(acc["email"], acc["password"], imap_server, msg.uid, folder)
        self._full_worker.finished_ok.connect(self._on_full_message_fetched)
        self._full_worker.finished_error.connect(
            lambda err: self.inbox_body.setPlainText(f"❌ Nie można wczytać treści: {err}")
        )
        self._full_worker.start()

    def _on_full_message_fetched(self, full_message):
        self._current_full_message = full_message
        if full_message.is_html:
            self.inbox_body.setHtml(full_message.body)
        else:
            self.inbox_body.setPlainText(full_message.body)
        self.inbox_attachments_list.clear()
        for att in full_message.attachments:
            size_kb = max(1, att.size // 1024)
            item = QListWidgetItem(f"📎 {att.filename} ({size_kb} KB)")
            item.setData(Qt.UserRole, att)
            self.inbox_attachments_list.addItem(item)

    def _save_inbox_attachment(self, item):
        att = item.data(Qt.UserRole)
        if not att:
            return
        path, _ = QFileDialog.getSaveFileName(self, "Zapisz załącznik", att.filename)
        if not path:
            return
        try:
            with open(path, "wb") as f:
                f.write(att.data)
            QMessageBox.information(self, tr('OK'), f"Zapisano: {path}")
        except OSError as e:
            QMessageBox.warning(self, tr('Błąd'), f"Nie udało się zapisać pliku: {e}")

    def _toggle_read_inbox_message(self):
        msg = self._selected_inbox_message()
        acc = self._current_inbox_account()
        if not msg or not acc:
            return
        imap_server = self.inbox_server_edit.text().strip()
        folder = self.inbox_folder_combo.currentText().strip() or "INBOX"
        new_seen_state = not msg.seen

        self._action_worker = MessageActionWorker(
            "flag", acc["email"], acc["password"], imap_server,
            msg.uid, folder, flag="\\Seen", add=new_seen_state
        )
        self._action_worker.finished_ok.connect(self.refresh_inbox)
        self._action_worker.finished_error.connect(
            lambda err: QMessageBox.warning(self, tr('Błąd'), f"Nie udało się zmienić statusu: {err}")
        )
        self._action_worker.start()

    def _delete_inbox_message(self):
        msg = self._selected_inbox_message()
        acc = self._current_inbox_account()
        if not msg or not acc:
            return
        reply = QMessageBox.question(
            self, tr('Potwierdź'), f"Usunąć wiadomość „{msg.subject}”?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.No:
            return

        imap_server = self.inbox_server_edit.text().strip()
        folder = self.inbox_folder_combo.currentText().strip() or "INBOX"
        trash = guess_trash_folder(getattr(self, "_inbox_folders", []))

        self._action_worker = MessageActionWorker(
            "delete", acc["email"], acc["password"], imap_server,
            msg.uid, folder, trash_folder=trash
        )
        self._action_worker.finished_ok.connect(self.refresh_inbox)
        self._action_worker.finished_error.connect(
            lambda err: QMessageBox.warning(self, tr('Błąd'), f"Nie udało się usunąć wiadomości: {err}")
        )
        self._action_worker.start()

    def _reply_to_inbox_message(self):
        msg = self._selected_inbox_message()
        acc = self._current_inbox_account()
        if not msg or not acc:
            QMessageBox.warning(self, tr('Błąd'), tr('Wybierz najpierw wiadomość.'))
            return
        full = self._current_full_message
        to_addr = full.sender_email if full else msg.sender
        original_body = full.body if full else ""
        quoted = "\n".join(f"> {line}" for line in original_body.splitlines()[:20])

        dialog = QDialog(self)
        dialog.setWindowTitle(tr('✉️ Odpowiedz'))
        dialog.resize(500, 400)
        vbox = QVBoxLayout(dialog)

        form = QFormLayout()
        to_edit = QLineEdit(to_addr)
        form.addRow(tr('Do:'), to_edit)
        subject_text = msg.subject if msg.subject.lower().startswith("re:") else f"Re: {msg.subject}"
        subject_edit = QLineEdit(subject_text)
        form.addRow(tr('Temat:'), subject_edit)
        vbox.addLayout(form)

        body_edit = QTextEdit()
        body_edit.setPlainText(f"\n\n---\n{quoted}")
        vbox.addWidget(body_edit)

        btn_send = QPushButton(tr('📤 Wyślij odpowiedź'))
        vbox.addWidget(btn_send)

        def _send():
            ok, message, _ = wyslij_email(
                to_edit.text().strip(), subject_edit.text().strip(), body_edit.toPlainText(),
                acc["email"], acc["password"],
                host=acc.get("host") or SMTP_RELAY_HOST,
                check_blacklist=False,
            )
            if ok:
                QMessageBox.information(dialog, tr('OK'), tr('Odpowiedź wysłana!'))
                dialog.accept()
            else:
                QMessageBox.warning(dialog, tr('Błąd'), f"Nie udało się wysłać: {message}")

        btn_send.clicked.connect(_send)
        dialog.exec()

    def refresh_history(self):
        rows = db.get_history(limit=200)
        self.history_table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val else "")
                if j == 2:
                    item.setBackground(QColor(*COLOR_OK) if val == 'wysłano' else QColor(*COLOR_ERROR))
                self.history_table.setItem(i, j, item)
        self.history_table.resizeColumnsToContents()

    def clear_old_logs(self):
        reply = QMessageBox.question(
            self, tr('Potwierdzenie'), tr('Usunąć logi starsze niż 30 dni?'),
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            removed = db.clear_old_logs(days=30)
            self.refresh_history()
            QMessageBox.information(self, tr('OK'), f"Usunięto {removed} starych wpisów.")

    def refresh_stats(self):
        if not MATPLOTLIB_AVAILABLE:
            QMessageBox.warning(self, tr('Błąd'), tr('Matplotlib nie jest zainstalowany.'))
            return
        from datetime import datetime, timedelta
        end = datetime.now()
        start = end - timedelta(days=30)
        with db.get_connection_context() as conn:
            rows = conn.execute(
                "SELECT date(data_wyslania) as day, COUNT(*) as cnt FROM wysylki "
                "WHERE data_wyslania >= ? AND status='wysłano' "
                "GROUP BY day ORDER BY day",
                (start.isoformat(),)
            ).fetchall()
        if not rows:
            QMessageBox.information(self, tr('Info'), tr('Brak danych.'))
            return

        days = [r[0] for r in rows]
        counts = [r[1] for r in rows]

        fig = Figure(figsize=(6, 3), dpi=100)
        ax = fig.add_subplot(111)
        ax.bar(days, counts, color='#89b4fa')
        ax.set_title('Wysyłki dzienne')
        ax.set_xlabel('Data')
        ax.set_ylabel('Liczba')
        fig.autofmt_xdate()

        canvas = FigureCanvasQTAgg(fig)
        if self.stats_canvas:
            self.stats_frame.layout().replaceWidget(self.stats_canvas, canvas)
            self.stats_canvas.deleteLater()
        else:
            self.stats_frame.layout().insertWidget(1, canvas)
        self.stats_canvas = canvas

    # ------------------------------------------------------------------
    # ZAKŁADKA: AI ASSISTANT
    # ------------------------------------------------------------------
    def build_ai_tab(self):
        container = QWidget()
        layout = QVBoxLayout(container)

        ai_intro_top = QLabel(
            tr('🤖 Asystent AI to opcjonalny dodatek – pisze szablony, tematy i pomaga ocenić leady. Program działa też bez niego. Aby go włączyć, wybierz dostawcę poniżej i zapisz konfigurację.')
        )
        ai_intro_top.setWordWrap(True)
        ai_intro_top.setStyleSheet("color: #a6adc8; font-size: 12px;")
        layout.addWidget(ai_intro_top)

        # === Konfiguracja AI ===
        cfg_group = QGroupBox(tr('⚙️ Konfiguracja dostawcy AI'))
        cfg_layout = QVBoxLayout(cfg_group)

        row_provider = QHBoxLayout()
        row_provider.addWidget(QLabel(tr('Provider:')))
        self.ai_provider_combo = QComboBox()
        self.ai_provider_combo.addItems(["OpenAI (ChatGPT)", "Google Gemini", "Ollama (Local)", "LM Studio (Local)", "DeepSeekLaude (Local)"])
        self.ai_provider_combo.currentIndexChanged.connect(self._ai_provider_changed)
        row_provider.addWidget(self.ai_provider_combo)
        btn_test_ai = QPushButton(tr('Test polaczenia'))
        btn_test_ai.clicked.connect(self._ai_test_connection)
        row_provider.addWidget(btn_test_ai)
        row_provider.addStretch()
        cfg_layout.addLayout(row_provider)

        # OpenAI
        self.ai_cfg_openai = QWidget()
        f_openai = QFormLayout(self.ai_cfg_openai)
        self.ai_openai_key = QLineEdit()
        self.ai_openai_key.setEchoMode(QLineEdit.Password)
        self.ai_openai_key.setPlaceholderText(tr('sk-...'))
        f_openai.addRow(tr('API Key:'), self.ai_openai_key)
        self.ai_openai_model = QComboBox()
        self.ai_openai_model.addItems(["gpt-3.5-turbo", "gpt-4", "gpt-4-turbo"])
        f_openai.addRow(tr('Model:'), self.ai_openai_model)
        cfg_layout.addWidget(self.ai_cfg_openai)

        # Gemini
        self.ai_cfg_gemini = QWidget()
        f_gemini = QFormLayout(self.ai_cfg_gemini)
        self.ai_gemini_key = QLineEdit()
        self.ai_gemini_key.setEchoMode(QLineEdit.Password)
        self.ai_gemini_key.setPlaceholderText(tr('AIza...'))
        f_gemini.addRow(tr('API Key:'), self.ai_gemini_key)
        cfg_layout.addWidget(self.ai_cfg_gemini)
        self.ai_cfg_gemini.setVisible(False)

        # Ollama
        self.ai_cfg_ollama = QWidget()
        f_ollama = QFormLayout(self.ai_cfg_ollama)
        self.ai_ollama_url = QLineEdit("http://localhost:11434")
        f_ollama.addRow(tr('URL:'), self.ai_ollama_url)
        self.ai_ollama_model = QComboBox()
        self.ai_ollama_model.addItems(["llama2", "mistral", "neural-chat", "dolphin-mixtral"])
        self.ai_ollama_model.setEditable(True)
        f_ollama.addRow(tr('Model:'), self.ai_ollama_model)
        cfg_layout.addWidget(self.ai_cfg_ollama)
        self.ai_cfg_ollama.setVisible(False)

        # LM Studio
        self.ai_cfg_lmstudio = QWidget()
        f_lm = QFormLayout(self.ai_cfg_lmstudio)
        self.ai_lmstudio_url = QLineEdit("http://localhost:1234")
        f_lm.addRow(tr('URL:'), self.ai_lmstudio_url)
        cfg_layout.addWidget(self.ai_cfg_lmstudio)
        self.ai_cfg_lmstudio.setVisible(False)

        # DeepSeekLaude (bridge do darmowego czatu DeepSeek, kompatybilny z OpenAI)
        self.ai_cfg_deepseeklaude = QWidget()
        f_dsl = QFormLayout(self.ai_cfg_deepseeklaude)
        self.ai_deepseeklaude_url = QLineEdit("http://127.0.0.1:8000")
        f_dsl.addRow(tr('URL:'), self.ai_deepseeklaude_url)
        self.ai_deepseeklaude_model = QComboBox()
        self.ai_deepseeklaude_model.addItems(["deepseek-chat", "deepseek-expert"])
        f_dsl.addRow(tr('Model:'), self.ai_deepseeklaude_model)
        self.ai_deepseeklaude_search = QCheckBox(tr('Wlacz wyszukiwanie w internecie (DeepSeek Search)'))
        self.ai_deepseeklaude_search.setToolTip(
            tr("Ta sama funkcja co przycisk 'Search' w czacie chat.deepseek.com - model przeszukuje internet przed odpowiedzia. Wolniejsze, ale przydatne np. do sprawdzenia firmy/leada w sieci.")
        )
        f_dsl.addRow("", self.ai_deepseeklaude_search)
        cfg_layout.addWidget(self.ai_cfg_deepseeklaude)
        self.ai_cfg_deepseeklaude.setVisible(False)

        btn_save_ai_cfg = QPushButton(tr('Zapisz konfiguracje AI'))
        btn_save_ai_cfg.clicked.connect(self._ai_save_config)
        cfg_layout.addWidget(btn_save_ai_cfg)

        self.ai_status_label = QLabel(tr('Status: brak skonfigurowanego providera'))
        self.ai_status_label.setStyleSheet("color: #a6adc8; font-size: 11px;")
        cfg_layout.addWidget(self.ai_status_label)

        layout.addWidget(cfg_group)

        # === Funkcje AI ===
        ai_tabs = QTabWidget()

        # --- TAB: Szablony ---
        tab_templates = QWidget()
        tl = QVBoxLayout(tab_templates)
        tl.addWidget(QLabel(tr('Generator szablonow e-mail (3 warianty)')))
        row_t = QHBoxLayout()
        row_t.addWidget(QLabel(tr('Branza:')))
        self.ai_gen_industry = QLineEdit()
        self.ai_gen_industry.setPlaceholderText(tr('np. IT, Finance, Budownictwo'))
        row_t.addWidget(self.ai_gen_industry)
        row_t.addWidget(QLabel(tr('Produkt/Usluga:')))
        self.ai_gen_product = QLineEdit()
        self.ai_gen_product.setPlaceholderText(tr('np. Oprogramowanie CRM'))
        row_t.addWidget(self.ai_gen_product)
        tl.addLayout(row_t)
        btn_gen_tmpl = QPushButton(tr('Generuj 3 warianty szablonow'))
        btn_gen_tmpl.setStyleSheet("background: #2b5e2b; font-weight: bold; padding: 6px;")
        btn_gen_tmpl.clicked.connect(self._ai_generate_templates)
        tl.addWidget(btn_gen_tmpl)
        self.ai_templates_result = QTextEdit()
        self.ai_templates_result.setReadOnly(True)
        self.ai_templates_result.setPlaceholderText(tr('Wygenerowane szablony pojawia sie tutaj...'))
        tl.addWidget(self.ai_templates_result)
        btn_use_template = QPushButton(tr('Uzyj pierwszego wariantu w zakladce Wysylka'))
        btn_use_template.clicked.connect(self._ai_use_first_template)
        tl.addWidget(btn_use_template)
        ai_tabs.addTab(tab_templates, tr('✉️ Szablony maili'))

        # --- TAB: Subject Lines ---
        tab_subject = QWidget()
        sl = QVBoxLayout(tab_subject)
        sl.addWidget(QLabel(tr('Optymalizator subject lines')))
        row_s = QHBoxLayout()
        row_s.addWidget(QLabel(tr('Temat:')))
        self.ai_subj_topic = QLineEdit()
        self.ai_subj_topic.setPlaceholderText(tr('np. Zwiekszenie sprzedazy'))
        row_s.addWidget(self.ai_subj_topic)
        row_s.addWidget(QLabel(tr('Branza:')))
        self.ai_subj_industry = QLineEdit()
        self.ai_subj_industry.setPlaceholderText(tr('np. IT'))
        row_s.addWidget(self.ai_subj_industry)
        sl.addLayout(row_s)
        btn_gen_subj = QPushButton(tr('Generuj 5 wariantow subject line'))
        btn_gen_subj.setStyleSheet("background: #2b5e2b; font-weight: bold; padding: 6px;")
        btn_gen_subj.clicked.connect(self._ai_generate_subjects)
        sl.addWidget(btn_gen_subj)
        self.ai_subj_list = QListWidget()
        sl.addWidget(self.ai_subj_list)
        row_score = QHBoxLayout()
        row_score.addWidget(QLabel(tr('Oceń subject:')))
        self.ai_subj_to_score = QLineEdit()
        self.ai_subj_to_score.setPlaceholderText(tr('Wklej subject line do oceny...'))
        row_score.addWidget(self.ai_subj_to_score)
        btn_score_subj = QPushButton(tr('Ocen (1-100)'))
        btn_score_subj.clicked.connect(self._ai_score_subject)
        row_score.addWidget(btn_score_subj)
        self.ai_subj_score_result = QLabel("")
        self.ai_subj_score_result.setStyleSheet("font-weight: bold; font-size: 14px;")
        row_score.addWidget(self.ai_subj_score_result)
        sl.addLayout(row_score)
        btn_use_subj = QPushButton(tr('Uzyj zaznaczonego w zakladce Wysylka'))
        btn_use_subj.clicked.connect(self._ai_use_selected_subject)
        sl.addWidget(btn_use_subj)
        ai_tabs.addTab(tab_subject, tr('📝 Tematy wiadomości'))

        # --- TAB: Lead Scoring ---
        tab_scoring = QWidget()
        scl = QVBoxLayout(tab_scoring)
        scl.addWidget(QLabel(tr('Ocenianie jakosci leadow (AI Quality Check)')))
        row_sc = QHBoxLayout()
        self.btn_score_all = QPushButton(tr('Ocen WSZYSTKIE leady'))
        self.btn_score_all.setStyleSheet("background: #5e2b8b; font-weight: bold; padding: 6px;")
        self.btn_score_all.clicked.connect(self._ai_score_all_leads)
        row_sc.addWidget(self.btn_score_all)
        self.btn_score_sel = QPushButton(tr('Ocen zaznaczone leady'))
        self.btn_score_sel.clicked.connect(self._ai_score_selected_leads)
        row_sc.addWidget(self.btn_score_sel)
        self.stop_scoring_btn = QPushButton(tr('⏹️ STOP'))
        self.stop_scoring_btn.setStyleSheet("background: #8b0000; font-weight: bold;")
        self.stop_scoring_btn.clicked.connect(self.stop_scoring)
        self.stop_scoring_btn.setEnabled(False)
        row_sc.addWidget(self.stop_scoring_btn)
        scl.addLayout(row_sc)
        self.ai_scoring_progress = QProgressBar()
        scl.addWidget(self.ai_scoring_progress)
        self.ai_scoring_table = QTableWidget()
        self.ai_scoring_table.setColumnCount(5)
        self.ai_scoring_table.setHorizontalHeaderLabels([tr('Email'), tr('Score'), tr('Spam?'), tr('Powod'), tr('Akcja')])
        self.ai_scoring_table.horizontalHeader().setStretchLastSection(True)
        scl.addWidget(self.ai_scoring_table)
        ai_tabs.addTab(tab_scoring, tr('⭐ Ocena leadów'))

        # --- TAB: Personalizacja ---
        tab_pers = QWidget()
        pl = QVBoxLayout(tab_pers)
        pl.addWidget(QLabel(tr('Personalizacja wiadomosci na podstawie strony www firmy')))
        pl.addWidget(QLabel(tr('Wklej adres email i website jednego leada:')))
        row_p = QHBoxLayout()
        row_p.addWidget(QLabel(tr('Email:')))
        self.ai_pers_email = QLineEdit()
        row_p.addWidget(self.ai_pers_email)
        row_p.addWidget(QLabel(tr('Website:')))
        self.ai_pers_website = QLineEdit()
        row_p.addWidget(self.ai_pers_website)
        row_p.addWidget(QLabel(tr('Firma:')))
        self.ai_pers_firma = QLineEdit()
        row_p.addWidget(self.ai_pers_firma)
        pl.addLayout(row_p)
        btn_analyze = QPushButton(tr('Analizuj website i personalizuj wiadomosc'))
        btn_analyze.setStyleSheet("background: #2b5e2b; font-weight: bold; padding: 6px;")
        btn_analyze.clicked.connect(self._ai_personalize_lead)
        pl.addWidget(btn_analyze)
        self.ai_pers_result = QTextEdit()
        self.ai_pers_result.setReadOnly(True)
        self.ai_pers_result.setPlaceholderText(tr('Spersonalizowana wiadomosc pojawi sie tutaj...'))
        pl.addWidget(self.ai_pers_result)
        ai_tabs.addTab(tab_pers, tr('👤 Personalizacja'))

        # --- TAB: Analiza odpowiedzi ---
        tab_resp = QWidget()
        rl = QVBoxLayout(tab_resp)
        rl.addWidget(QLabel(tr('Analizator odpowiedzi na e-maile')))
        rl.addWidget(QLabel(tr('Wklej otrzymana odpowiedz:')))
        self.ai_resp_input = QTextEdit()
        self.ai_resp_input.setMaximumHeight(150)
        self.ai_resp_input.setPlaceholderText(tr('Wklej tresc otrzymanej wiadomosci...'))
        rl.addWidget(self.ai_resp_input)
        btn_analyze_resp = QPushButton(tr('Analizuj odpowiedz'))
        btn_analyze_resp.setStyleSheet("background: #2b5e2b; font-weight: bold; padding: 6px;")
        btn_analyze_resp.clicked.connect(self._ai_analyze_response)
        rl.addWidget(btn_analyze_resp)
        self.ai_resp_result = QTextEdit()
        self.ai_resp_result.setReadOnly(True)
        self.ai_resp_result.setPlaceholderText(tr('Wynik analizy pojawi sie tutaj...'))
        rl.addWidget(self.ai_resp_result)
        ai_tabs.addTab(tab_resp, tr('📨 Analiza odpowiedzi'))

        # --- TAB: Timing ---
        tab_timing = QWidget()
        til = QVBoxLayout(tab_timing)
        til.addWidget(QLabel(tr('Optymalizacja czasu wysylki')))
        row_ti = QHBoxLayout()
        row_ti.addWidget(QLabel(tr('Branza:')))
        self.ai_timing_industry = QLineEdit()
        self.ai_timing_industry.setPlaceholderText(tr('np. IT, Finance'))
        row_ti.addWidget(self.ai_timing_industry)
        row_ti.addWidget(QLabel(tr('Region:')))
        self.ai_timing_region = QComboBox()
        self.ai_timing_region.addItems(["DE", "PL", "US", "GB", "FR", "AT", "CH"])
        row_ti.addWidget(self.ai_timing_region)
        row_ti.addWidget(QLabel(tr('Rola:')))
        self.ai_timing_role = QLineEdit()
        self.ai_timing_role.setPlaceholderText(tr('np. manager, CEO'))
        self.ai_timing_role.setText(tr('manager'))
        row_ti.addWidget(self.ai_timing_role)
        til.addLayout(row_ti)
        btn_get_timing = QPushButton(tr('Oblicz optymalny czas wysylki'))
        btn_get_timing.setStyleSheet("background: #2b5e2b; font-weight: bold; padding: 6px;")
        btn_get_timing.clicked.connect(self._ai_get_timing)
        til.addWidget(btn_get_timing)
        self.ai_timing_result = QTextEdit()
        self.ai_timing_result.setReadOnly(True)
        til.addWidget(self.ai_timing_result)
        ai_tabs.addTab(tab_timing, tr('⏰ Najlepszy czas wysyłki'))

        # --- TAB: A/B Testing ---
        tab_ab = QWidget()
        abl = QVBoxLayout(tab_ab)
        abl.addWidget(QLabel(tr('Generator wariantow A/B')))
        row_ab = QHBoxLayout()
        row_ab.addWidget(QLabel(tr('Typ:')))
        self.ai_ab_type = QComboBox()
        self.ai_ab_type.addItems(["subject_line", "email_body", "cta"])
        row_ab.addWidget(self.ai_ab_type)
        row_ab.addStretch()
        abl.addLayout(row_ab)
        abl.addWidget(QLabel(tr('Oryginal:')))
        self.ai_ab_original = QTextEdit()
        self.ai_ab_original.setMaximumHeight(100)
        self.ai_ab_original.setPlaceholderText(tr('Wklej oryginalny subject line lub tresc...'))
        abl.addWidget(self.ai_ab_original)
        btn_gen_ab = QPushButton(tr('Generuj 2 warianty A/B'))
        btn_gen_ab.setStyleSheet("background: #2b5e2b; font-weight: bold; padding: 6px;")
        btn_gen_ab.clicked.connect(self._ai_generate_ab)
        abl.addWidget(btn_gen_ab)
        self.ai_ab_result = QTextEdit()
        self.ai_ab_result.setReadOnly(True)
        self.ai_ab_result.setPlaceholderText(tr('Warianty A/B pojawia sie tutaj...'))
        abl.addWidget(self.ai_ab_result)
        ai_tabs.addTab(tab_ab, tr('🔀 Testy A/B'))

        layout.addWidget(ai_tabs)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(container)
        self.ai_tab.layout().addWidget(scroll)

    # ------------------------------------------------------------------
    # AI - KONFIGURACJA
    # ------------------------------------------------------------------
    def _ai_provider_changed(self, index):
        self.ai_cfg_openai.setVisible(index == 0)
        self.ai_cfg_gemini.setVisible(index == 1)
        self.ai_cfg_ollama.setVisible(index == 2)
        self.ai_cfg_lmstudio.setVisible(index == 3)
        self.ai_cfg_deepseeklaude.setVisible(index == 4)

    def _load_ai_config(self):
        """Wczytuje zapisaną konfigurację AI (provider, klucze, model) z ustawień
        profilu z powrotem do UI i ponownie aktywuje providera w ai_manager.
        Wcześniej te ustawienia były tylko zapisywane (db.set_setting), ale nigdy
        nie odczytywane z powrotem - po restarcie aplikacji lub zmianie profilu
        wybór providera/modelu/klucza API wyglądał jakby "nie zapamiętał się",
        mimo że w bazie danych był poprawnie zapisany."""
        provider_name = db.get_setting("ai_provider", "")
        index_by_name = {"openai": 0, "gemini": 1, "ollama": 2, "lmstudio": 3, "deepseeklaude": 4}

        self.ai_openai_key.setText(db.get_setting("ai_openai_key", ""))
        saved_openai_model = db.get_setting("ai_openai_model", "gpt-3.5-turbo")
        idx = self.ai_openai_model.findText(saved_openai_model)
        self.ai_openai_model.setCurrentIndex(idx if idx >= 0 else 0)

        self.ai_gemini_key.setText(db.get_setting("ai_gemini_key", ""))

        self.ai_ollama_url.setText(db.get_setting("ai_ollama_url", "http://localhost:11434"))
        saved_ollama_model = db.get_setting("ai_ollama_model", "")
        if saved_ollama_model:
            idx = self.ai_ollama_model.findText(saved_ollama_model)
            if idx >= 0:
                self.ai_ollama_model.setCurrentIndex(idx)
            else:
                self.ai_ollama_model.setCurrentText(saved_ollama_model)

        self.ai_lmstudio_url.setText(db.get_setting("ai_lmstudio_url", "http://localhost:1234"))

        self.ai_deepseeklaude_url.setText(db.get_setting("ai_deepseeklaude_url", "http://127.0.0.1:8000"))
        saved_dsl_model = db.get_setting("ai_deepseeklaude_model", "deepseek-chat")
        idx = self.ai_deepseeklaude_model.findText(saved_dsl_model)
        self.ai_deepseeklaude_model.setCurrentIndex(idx if idx >= 0 else 0)
        self.ai_deepseeklaude_search.setChecked(db.get_setting("ai_deepseeklaude_search", "0") == "1")

        if provider_name not in index_by_name:
            self.ai_status_label.setText(tr('Status: brak skonfigurowanego providera'))
            return

        self.ai_provider_combo.blockSignals(True)
        self.ai_provider_combo.setCurrentIndex(index_by_name[provider_name])
        self.ai_provider_combo.blockSignals(False)
        self._ai_provider_changed(index_by_name[provider_name])

        try:
            if provider_name == "openai":
                provider = OpenAIProvider(self.ai_openai_key.text().strip(), self.ai_openai_model.currentText())
            elif provider_name == "gemini":
                provider = GeminiProvider(self.ai_gemini_key.text().strip())
            elif provider_name == "ollama":
                provider = OllamaProvider(self.ai_ollama_url.text().strip(), self.ai_ollama_model.currentText())
            elif provider_name == "lmstudio":
                provider = LMStudioProvider(self.ai_lmstudio_url.text().strip())
            else:
                provider = DeepSeekLaudeProvider(
                    self.ai_deepseeklaude_url.text().strip(),
                    self.ai_deepseeklaude_model.currentText(),
                    self.ai_deepseeklaude_search.isChecked(),
                )
            ai_manager.register_provider(provider_name, provider)
            ai_manager.set_active_provider(provider_name)
            self.ai_status_label.setText(f"Status: Provider wczytany - {self.ai_provider_combo.currentText()}")
            self.ai_status_label.setStyleSheet("color: #a6e3a1;")
        except Exception as e:
            logger.warning("Nie udało się przywrócić konfiguracji AI: %s", e)

    def _ai_save_config(self):
        index = self.ai_provider_combo.currentIndex()
        try:
            if index == 0:
                key = self.ai_openai_key.text().strip()
                model = self.ai_openai_model.currentText()
                provider = OpenAIProvider(key, model)
                ai_manager.register_provider("openai", provider)
                ai_manager.set_active_provider("openai")
                db.set_setting("ai_provider", "openai")
                db.set_setting("ai_openai_key", key)
                db.set_setting("ai_openai_model", model)
            elif index == 1:
                key = self.ai_gemini_key.text().strip()
                provider = GeminiProvider(key)
                ai_manager.register_provider("gemini", provider)
                ai_manager.set_active_provider("gemini")
                db.set_setting("ai_provider", "gemini")
                db.set_setting("ai_gemini_key", key)
            elif index == 2:
                url = self.ai_ollama_url.text().strip()
                model = self.ai_ollama_model.currentText()
                provider = OllamaProvider(url, model)
                ai_manager.register_provider("ollama", provider)
                ai_manager.set_active_provider("ollama")
                db.set_setting("ai_provider", "ollama")
                db.set_setting("ai_ollama_url", url)
                db.set_setting("ai_ollama_model", model)
            elif index == 3:
                url = self.ai_lmstudio_url.text().strip()
                provider = LMStudioProvider(url)
                ai_manager.register_provider("lmstudio", provider)
                ai_manager.set_active_provider("lmstudio")
                db.set_setting("ai_provider", "lmstudio")
                db.set_setting("ai_lmstudio_url", url)
            elif index == 4:
                url = self.ai_deepseeklaude_url.text().strip()
                model = self.ai_deepseeklaude_model.currentText()
                web_search = self.ai_deepseeklaude_search.isChecked()
                provider = DeepSeekLaudeProvider(url, model, web_search)
                ai_manager.register_provider("deepseeklaude", provider)
                ai_manager.set_active_provider("deepseeklaude")
                db.set_setting("ai_provider", "deepseeklaude")
                db.set_setting("ai_deepseeklaude_url", url)
                db.set_setting("ai_deepseeklaude_model", model)
                db.set_setting("ai_deepseeklaude_search", "1" if web_search else "0")
            self.ai_status_label.setText(f"Status: Provider zapisany - {self.ai_provider_combo.currentText()}")
            self.ai_status_label.setStyleSheet("color: #a6e3a1;")
            QMessageBox.information(self, tr('OK'), tr('Konfiguracja AI zapisana!'))
        except Exception as e:
            QMessageBox.critical(self, tr('Blad'), f"Blad konfiguracji: {e}")

    def _ai_test_connection(self):
        provider = ai_manager.get_active_provider()
        if not provider:
            QMessageBox.warning(self, tr('Blad'), tr('Najpierw zapisz konfiguracje AI!'))
            return
        self.ai_status_label.setText(tr('Testuje polaczenie...'))
        if provider.check_connection():
            self.ai_status_label.setText(f"Status: Polaczono z {provider.name}")
            self.ai_status_label.setStyleSheet("color: #a6e3a1;")
            QMessageBox.information(self, tr('OK'), f"Polaczenie z {provider.name} dziala!")
        else:
            self.ai_status_label.setText(f"Status: Blad polaczenia z {provider.name}")
            self.ai_status_label.setStyleSheet("color: #f38ba8;")
            QMessageBox.warning(self, tr('Blad'), f"Nie mozna polaczyc sie z {provider.name}")

    def _ai_check_provider(self) -> bool:
        if not ai_manager.get_active_provider():
            QMessageBox.warning(
                self, tr('AI nie jest skonfigurowane'),
                tr('Aby korzystać z Asystenta AI, ustaw dostawcę AI w sekcji "Konfiguracja AI"\nna górze zakładki: 📁 Kampania → 🤖 Asystent AI\n\nReszta programu (szukanie, leady, wysyłka) działa normalnie także bez AI.')
            )
            return False
        return True

    def _goto_send_tab(self):
        """Przełącza na zakładkę Wysyłka -> Wyślij kampanię."""
        self.tabs.setCurrentIndex(1)
        self.wysylka_tabs.setCurrentIndex(0)

    def _on_language_changed(self, index):
        """Zapisuje wybrany język na dysku i od razu proponuje restart.
        Nie próbujemy przepisywać całego GUI "na żywo" - to byłoby bardzo
        ryzykowne w tak dużym, ręcznie budowanym interfejsie. Restart jest
        prosty i pewny."""
        code = self.language_combo.itemData(index)
        if not code or code == get_language():
            return
        set_language(code)
        reply = QMessageBox.question(
            self, tr("Zrestartować teraz?"),
            tr("Język został zmieniony. Aby zastosować zmiany, aplikacja musi zostać zrestartowana teraz."),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if reply == QMessageBox.Yes:
            restart_app()

    # ------------------------------------------------------------------
    # AI - SZABLONY
    # ------------------------------------------------------------------
    def _ai_generate_templates(self):
        if not self._ai_check_provider():
            return
        industry = self.ai_gen_industry.text().strip()
        product = self.ai_gen_product.text().strip()
        if not industry or not product:
            QMessageBox.warning(self, tr('Blad'), tr('Wpisz branze i produkt!'))
            return
        self.ai_status_label.setText(tr('Generuje szablony...'))
        self.ai_templates_result.setPlainText(tr('Proszę czekac...'))
        templates = TemplateGenerator.generate(industry, product)
        if templates:
            result = "\n\n" + "=" * 60 + "\n\n".join(
                [f"WARIANT {i+1}:\n{t}" for i, t in enumerate(templates)]
            )
            self.ai_templates_result.setPlainText(result)
            self.ai_status_label.setText(f"Wygenerowano {len(templates)} wariantow")
            self.ai_status_label.setStyleSheet("color: #a6e3a1;")
        else:
            self.ai_templates_result.setPlainText(tr('Blad generowania - sprawdz konfiguracje AI'))
            self.ai_status_label.setText(tr('Blad generowania'))
            self.ai_status_label.setStyleSheet("color: #f38ba8;")

    def _ai_use_first_template(self):
        text = self.ai_templates_result.toPlainText()
        if not text or text == tr("Proszę czekac..."):
            QMessageBox.warning(self, tr('Blad'), tr('Najpierw wygeneruj szablony!'))
            return
        parts = text.split("=" * 60)
        if len(parts) >= 2:
            first = parts[1].strip()
            if first.startswith("WARIANT 1:"):
                first = first[len("WARIANT 1:"):].strip()
            self.szablon_edit.setPlainText(first)
            self._goto_send_tab()
            QMessageBox.information(self, tr('OK'), tr('Szablon wklejony do zakładki Wysyłka!'))

    # ------------------------------------------------------------------
    # AI - SUBJECT LINES
    # ------------------------------------------------------------------
    def _ai_generate_subjects(self):
        if not self._ai_check_provider():
            return
        topic = self.ai_subj_topic.text().strip()
        industry = self.ai_subj_industry.text().strip()
        if not topic or not industry:
            QMessageBox.warning(self, tr('Blad'), tr('Wpisz temat i branze!'))
            return
        self.ai_status_label.setText(tr('Generuje subject lines...'))
        self.ai_subj_list.clear()
        variants = SubjectLineOptimizer.generate_variants(topic, industry, 5)
        if variants:
            for v in variants:
                self.ai_subj_list.addItem(v)
            self.ai_status_label.setText(f"Wygenerowano {len(variants)} wariantow")
            self.ai_status_label.setStyleSheet("color: #a6e3a1;")
        else:
            self.ai_status_label.setText(tr('Blad generowania'))
            self.ai_status_label.setStyleSheet("color: #f38ba8;")

    def _ai_score_subject(self):
        if not self._ai_check_provider():
            return
        subject = self.ai_subj_to_score.text().strip()
        if not subject:
            QMessageBox.warning(self, tr('Blad'), tr('Wpisz subject line do oceny!'))
            return
        score = SubjectLineOptimizer.score_subject_line(subject)
        if score is not None:
            color = "#a6e3a1" if score > 70 else "#fab387" if score > 50 else "#f38ba8"
            self.ai_subj_score_result.setText(f"{score}/100")
            self.ai_subj_score_result.setStyleSheet(f"color: {color}; font-weight: bold; font-size: 16px;")
        else:
            self.ai_subj_score_result.setText(tr('Blad'))

    def _ai_use_selected_subject(self):
        item = self.ai_subj_list.currentItem()
        if not item:
            QMessageBox.warning(self, tr('Blad'), tr('Zaznacz subject line z listy!'))
            return
        self.temat_edit.setText(item.text())
        self._goto_send_tab()
        QMessageBox.information(self, tr('OK'), tr('Temat wiadomości wklejony do zakładki Wysyłka!'))

    # ------------------------------------------------------------------
    # AI - LEAD SCORING
    # ------------------------------------------------------------------
    def _ai_score_all_leads(self):
        if not self._ai_check_provider():
            return
        leads = db.get_leads()
        if not leads:
            QMessageBox.warning(self, tr('Blad'), tr('Brak leadow w bazie!'))
            return
        leads_list = [{"id": r[0], "firma": r[1] or "", "email": r[3] or "", "website": r[6] or ""} for r in leads if r[3]]
        self._ai_start_scoring(leads_list)

    def _ai_score_selected_leads(self):
        if not self._ai_check_provider():
            return
        rows = set()
        for item in self.leads_table.selectedItems():
            rows.add(item.row())
        if not rows:
            QMessageBox.warning(self, tr('Blad'), tr('Zaznacz leady w zakladce Leadzy!'))
            return
        leads_list = []
        for row in rows:
            email_item = self.leads_table.item(row, 2)
            firma_item = self.leads_table.item(row, 1)
            id_item = self.leads_table.item(row, 0)
            if email_item and email_item.text():
                leads_list.append({
                    "id": int(id_item.text()) if id_item else 0,
                    "firma": firma_item.text() if firma_item else "",
                    "email": email_item.text(),
                    "website": ""
                })
        self._ai_start_scoring(leads_list)

    def _ai_start_scoring(self, leads_list):
        industry = self.ai_timing_industry.text() or "general"
        self.ai_scoring_table.setRowCount(0)
        self.ai_scoring_progress.setValue(0)
        self.batch_ai_worker = BatchAIWorker(leads_list, "score", industry=industry)
        self.batch_ai_worker.progress.connect(lambda c, t: self.ai_scoring_progress.setValue(int(c * 100 / t)))
        self.batch_ai_worker.lead_result.connect(self._ai_on_lead_scored)
        self.batch_ai_worker.finished.connect(self._ai_on_scoring_finished)
        self.batch_ai_worker.start()
        self.btn_score_all.setEnabled(False)
        self.btn_score_sel.setEnabled(False)
        self.stop_scoring_btn.setEnabled(True)
        self.ai_status_label.setText(tr('Oceniam leady...'))
        self.ai_status_label.setStyleSheet("color: #89b4fa;")

    def stop_scoring(self):
        if self.batch_ai_worker:
            self.batch_ai_worker.stop()
            self.ai_status_label.setText(tr('⏹️ Zatrzymywanie...'))

    def _ai_on_scoring_finished(self):
        self.ai_status_label.setText(f"Oceniono {self.ai_scoring_table.rowCount()} leadow")
        self.ai_status_label.setStyleSheet("color: #a6e3a1;")
        self.btn_score_all.setEnabled(True)
        self.btn_score_sel.setEnabled(True)
        self.stop_scoring_btn.setEnabled(False)

    def _ai_on_lead_scored(self, email, result):
        # BatchAIWorker._score_lead() zwraca płaski słownik w "data"
        # (bez dodatkowego zagnieżdżenia pod kluczem "score_data") - stąd
        # kluczy szukamy bezpośrednio w result["data"].
        score_data = result.get("data") or {}
        if not score_data:
            return
        score = score_data.get("score", 0)
        is_spam = score_data.get("is_spam", False)
        reason = score_data.get("reason", "")
        action = score_data.get("recommended_action", "contact" if not is_spam else "skip")
        row = self.ai_scoring_table.rowCount()
        self.ai_scoring_table.insertRow(row)
        self.ai_scoring_table.setItem(row, 0, QTableWidgetItem(email))
        score_item = QTableWidgetItem(str(score))
        if score >= 70:
            score_item.setBackground(QColor(166, 227, 161, 80))
        elif score < 40:
            score_item.setBackground(QColor(243, 139, 168, 80))
        self.ai_scoring_table.setItem(row, 1, score_item)
        self.ai_scoring_table.setItem(row, 2, QTableWidgetItem("TAK" if is_spam else "NIE"))
        self.ai_scoring_table.setItem(row, 3, QTableWidgetItem(reason[:80] if reason else ""))
        self.ai_scoring_table.setItem(row, 4, QTableWidgetItem(action))

    # ------------------------------------------------------------------
    # AI - PERSONALIZACJA
    # ------------------------------------------------------------------
    def _ai_personalize_lead(self):
        if not self._ai_check_provider():
            return
        email = self.ai_pers_email.text().strip()
        website = self.ai_pers_website.text().strip()
        firma = self.ai_pers_firma.text().strip()
        template = self.szablon_edit.toPlainText().strip()
        if not email:
            QMessageBox.warning(self, tr('Blad'), tr('Wpisz adres email leadу!'))
            return
        if not template:
            QMessageBox.warning(self, tr('Blad'), tr('Najpierw wpisz szablon w zakladce Wysylka!'))
            return
        self.ai_status_label.setText(tr('Analizuje website i personalizuje...'))
        self.ai_pers_result.setPlainText(tr('Proszę czekac...'))
        industry = self.ai_timing_industry.text() or "general"
        insights = None
        if website:
            insights = LeadPersonalizer.analyze_website(firma, website, industry)
        lead = {"firma": firma, "email": email, "website": website}
        result = LeadPersonalizer.personalize_message(template, lead, insights)
        if result:
            self.ai_pers_result.setPlainText(result)
            self.ai_status_label.setText(tr('Personalizacja gotowa'))
            self.ai_status_label.setStyleSheet("color: #a6e3a1;")
        else:
            self.ai_pers_result.setPlainText(tr('Blad personalizacji - sprawdz konfiguracje AI'))
            self.ai_status_label.setText(tr('Blad personalizacji'))
            self.ai_status_label.setStyleSheet("color: #f38ba8;")

    # ------------------------------------------------------------------
    # AI - ANALIZA ODPOWIEDZI
    # ------------------------------------------------------------------
    def _ai_analyze_response(self):
        if not self._ai_check_provider():
            return
        body = self.ai_resp_input.toPlainText().strip()
        if not body:
            QMessageBox.warning(self, tr('Blad'), tr('Wklej tresc odpowiedzi!'))
            return
        self.ai_status_label.setText(tr('Analizuje odpowiedz...'))
        analysis = ResponseAnalyzer.classify_response(body)
        if analysis:
            typ = analysis.get("type", "unknown")
            sentiment = analysis.get("sentiment", "unknown")
            action = analysis.get("next_action", "unknown")
            points = ", ".join(analysis.get("key_points", []))
            color_map = {"interested": "#a6e3a1", "rejected": "#f38ba8", "more_info": "#fab387",
                         "spam": "#6c7086", "out_of_office": "#89b4fa"}
            color = color_map.get(typ, "#cdd6f4")
            result_text = f"Typ: {typ}\nSentiment: {sentiment}\nNastepny krok: {action}\nKluczowe punkty: {points}"
            self.ai_resp_result.setPlainText(result_text)
            self.ai_resp_result.setStyleSheet(f"color: {color};")
            self.ai_status_label.setText(f"Analiza: {typ} ({sentiment})")
            self.ai_status_label.setStyleSheet(f"color: {color};")
        else:
            self.ai_resp_result.setPlainText(tr('Blad analizy - sprawdz konfiguracje AI'))
            self.ai_status_label.setText(tr('Blad analizy'))
            self.ai_status_label.setStyleSheet("color: #f38ba8;")

    # ------------------------------------------------------------------
    # AI - TIMING
    # ------------------------------------------------------------------
    def _ai_get_timing(self):
        if not self._ai_check_provider():
            return
        industry = self.ai_timing_industry.text().strip() or "general"
        region = self.ai_timing_region.currentText()
        role = self.ai_timing_role.text().strip() or "manager"
        self.ai_status_label.setText(tr('Obliczam optymalny timing...'))
        timing = SendTimingOptimizer.recommend_send_time(industry, region, role)
        if timing:
            result = (
                f"Najlepszy dzien: {timing.get('best_day', 'N/A')}\n"
                f"Najlepszy czas: {timing.get('best_time', 'N/A')}\n"
                f"Strefa czasowa: {timing.get('best_timezone', 'N/A')}\n"
                f"Pewnosc: {timing.get('confidence', 0):.0%}\n"
                f"Uzasadnienie: {timing.get('reason', 'N/A')}"
            )
            self.ai_timing_result.setPlainText(result)
            self.ai_status_label.setText(tr('Rekomendacje gotowe'))
            self.ai_status_label.setStyleSheet("color: #a6e3a1;")
        else:
            self.ai_timing_result.setPlainText(tr('Blad - sprawdz konfiguracje AI'))
            self.ai_status_label.setText(tr('Blad'))
            self.ai_status_label.setStyleSheet("color: #f38ba8;")

    # ------------------------------------------------------------------
    # AI - A/B TESTING
    # ------------------------------------------------------------------
    def _ai_generate_ab(self):
        if not self._ai_check_provider():
            return
        original = self.ai_ab_original.toPlainText().strip()
        if not original:
            QMessageBox.warning(self, tr('Blad'), tr('Wpisz oryginal!'))
            return
        content_type = self.ai_ab_type.currentText()
        self.ai_status_label.setText(tr('Generuje warianty A/B...'))
        variants = ABTestingEngine.generate_variants(content_type, original, 2)
        if variants:
            result = f"ORYGINAL (A):\n{original}\n\n" + "=" * 50 + "\n\n"
            for i, v in enumerate(variants):
                result += f"WARIANT {chr(66+i)}:\n{v}\n\n"
            self.ai_ab_result.setPlainText(result)
            self.ai_status_label.setText(f"Wygenerowano {len(variants)} wariantow A/B")
            self.ai_status_label.setStyleSheet("color: #a6e3a1;")
        else:
            self.ai_ab_result.setPlainText(tr('Blad generowania - sprawdz konfiguracje AI'))
            self.ai_status_label.setText(tr('Blad'))
            self.ai_status_label.setStyleSheet("color: #f38ba8;")
    # AUTO-SAVE SZABLONOW I ZAPYTAN
    def _auto_save_template_subject(self):
        if not hasattr(self, 'szablon_edit') or not hasattr(self, 'temat_edit'):
            return
        settings = get_current_profile_settings()
        settings['last_template'] = self.szablon_edit.toPlainText()
        settings['last_subject'] = self.temat_edit.text()
        update_current_profile_settings(settings)

    def _manual_save_template(self):
        """Jawny, widoczny zapis szablonu (przycisk 'Zapisz szablon') - w
        odroznieniu od cichego auto-save (_auto_save_template_subject),
        daje uzytkownikowi wyrazne potwierdzenie ze zapis sie udal, zamiast
        polegac wylacznie na niewidocznym zapisie przy kazdym naciscnieciu
        klawisza."""
        try:
            settings = get_current_profile_settings()
            settings['last_template'] = self.szablon_edit.toPlainText()
            settings['last_subject'] = self.temat_edit.text()
            update_current_profile_settings(settings)
            QMessageBox.information(
                self, "Zapisano",
                "Szablon i temat zostały zapisane. Będą wczytane automatycznie "
                "przy następnym uruchomieniu programu (dla tego profilu)."
            )
        except Exception as e:
            QMessageBox.critical(self, "Błąd zapisu", f"Nie udało się zapisać szablonu:\n{e}")

    def _auto_save_search_params(self):
        if not hasattr(self, 'queries_edit') or not hasattr(self, 'locations_edit'):
            return
        settings = get_current_profile_settings()
        settings['last_queries'] = self.queries_edit.toPlainText()
        settings['last_locations'] = self.locations_edit.toPlainText()
        update_current_profile_settings(settings)

    def _on_ai_scoring_threshold_changed(self, value):
        if hasattr(self, 'lead_score_min_spin'):
            self.lead_score_min_spin.blockSignals(True)
            self.lead_score_min_spin.setValue(value)
            self.lead_score_min_spin.blockSignals(False)
        settings = get_current_profile_settings()
        settings['ai_scoring_threshold'] = value
        update_current_profile_settings(settings)

    def _auto_save_lead_score_threshold(self):
        if not hasattr(self, 'lead_score_min_spin') or not hasattr(self, 'lead_score_allow_unscored'):
            return
        settings = get_current_profile_settings()
        settings['ai_scoring_threshold'] = self.lead_score_min_spin.value()
        settings['lead_score_allow_unscored'] = self.lead_score_allow_unscored.isChecked()
        update_current_profile_settings(settings)
        # Zsynchronizuj też suwak w zakładce Ustawienia, jeśli już istnieje
        if hasattr(self, 'ai_scoring_threshold'):
            self.ai_scoring_threshold.blockSignals(True)
            self.ai_scoring_threshold.setValue(self.lead_score_min_spin.value())
            self.ai_scoring_threshold.blockSignals(False)
